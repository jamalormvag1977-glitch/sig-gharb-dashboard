#!/usr/bin/env python3
"""
Canevas Hebdomadaire V2 - Suivi d'Exécution Physique et Financière
ORMVAG - Région du Gharb
Avec indicateurs physiques par prestation (linéaire, volume, nombre...)
L'avancement physique % est calculé automatiquement à partir des quantités.
"""

import json
import sys
import os
import re

# ── Skill imports ──────────────────────────────────────────────
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers, Protection
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
from templates.base import (
    setup_sheet, style_header_row, style_data_row, style_total_row,
    font_title, font_header, font_subheader, font_body, font_caption, font_kpi, font_kpi_label,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_number, align_text, align_date,
    COLUMN_WIDTHS, FORMATS, ROW_HEIGHTS,
    PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_0,
    HEADER_BOLD, FONT_NAME, HEADER_TEXT,
    CF_POSITIVE_FILL, CF_POSITIVE_FONT, CF_NEGATIVE_FILL, CF_NEGATIVE_FONT,
    CF_WARNING_FILL, CF_WARNING_FONT,
    auto_fit_columns,
)

# Province colors from the GIS dashboard
PROVINCE_COLORS = {
    "Kénitra": "D4A017",
    "Sidi Kacem": "C76E7E",
    "Sidi Slimane": "5BB58A",
}
PROVINCE_FILLS = {
    "Kénitra": PatternFill("solid", fgColor="FFF8DC"),
    "Sidi Kacem": PatternFill("solid", fgColor="FFF0F3"),
    "Sidi Slimane": PatternFill("solid", fgColor="F0FFF4"),
}

# ── Load project data ──────────────────────────────────────────
DATA_PATH = "/tmp/sig-gharb-extract/src/data/dashboard_data.json"
with open(DATA_PATH, "r", encoding="utf-8") as f:
    data = json.load(f)
projects = data["projects"]
N = len(projects)

# ── Determine physical unit and estimated quantity per project ──
def infer_physical_indicator(consistance, rubrique, cout):
    """
    Infer the physical unit and an estimated planned quantity
    based on the consistance description and rubrique.
    Returns (unité, quantité_prévue_estimée)
    """
    c = (consistance or "").lower()
    r = (rubrique or "").lower()
    intitule = ""

    # ── Canal / Collecteur / Curage → ml (linéaire) ─────────
    if any(kw in c for kw in ["curage des collecteurs", "curage du canal", "curage des canaux",
                                "curage des émissaires", "curage des emissaires",
                                "curage du collecteur", "canal d'assainissement",
                                "canaux d'assainissement"]):
        # Estimate: ~500-3000 DH/ml for canal work
        qte = max(100, round(cout / 2000))
        return ("ml", qte)

    # ── Pistes agricoles → km ───────────────────────────────
    if any(kw in c for kw in ["pistes agricoles", "piste agricole"]):
        qte = max(1, round(cout / 500000))
        return ("km", qte)

    # ── Canaux portés → ml ──────────────────────────────────
    if any(kw in c for kw in ["canaux porté", "canaux porte", "recalage"]):
        qte = max(50, round(cout / 3000))
        return ("ml", qte)

    # ── Débitmètres → U (nombre d'équipements) ─────────────
    if "débitmètre" in c or "debitmetre" in c:
        # Count from description
        nums = re.findall(r'(\d+)\s*débitmètre', c)
        if nums:
            return ("U", int(nums[0]))
        if "sept" in c or "7" in c:
            return ("U", 7)
        if "trois" in c or "3" in c:
            return ("U", 3)
        if "quatre" in c or "4" in c:
            return ("U", 4)
        if "deux" in c or "2" in c:
            return ("U", 2)
        return ("U", 1)

    # ── Moteurs électriques → U ────────────────────────────
    if "moteur" in c:
        nums = re.findall(r'(\d+)\s*moteur', c)
        if nums:
            return ("U", int(nums[0]))
        if "trois" in c:
            return ("U", 3)
        if "deux" in c:
            return ("U", 2)
        return ("U", 1)

    # ── Pompes → U ─────────────────────────────────────────
    if "pompe" in c and "station" in c:
        nums = re.findall(r'(\d+)\s*pompe', c)
        if nums:
            return ("U", int(nums[0]))
        if "cinq" in c:
            return ("U", 5)
        return ("U", 1)

    # ── Gabions / Reno → m³ ────────────────────────────────
    if "gabion" in c or "reno" in c:
        qte = max(10, round(cout / 1500))
        return ("m³", qte)

    # ── Murs clôture → ml ──────────────────────────────────
    if "mur" in c and ("clôture" in c or "cloture" in c):
        # Extract meter count from description if available
        nums = re.findall(r'(\d+)\s*m', c)
        if nums:
            return ("ml", int(nums[0]))
        qte = max(20, round(cout / 3000))
        return ("ml", qte)

    # ── Étanchéité / Génie civil locaux → m² ──────────────
    if "étanchéité" in c or "etanchéité" in c or "etanchéite" in c or "locaux" in c:
        qte = max(20, round(cout / 2000))
        return ("m²", qte)

    # ── Ouvrages génie civil → U ──────────────────────────
    if "ouvrage" in c or "génie civil" in c or "genie civil" in c:
        qte = max(1, round(cout / 500000))
        return ("U", qte)

    # ── Bornes → U ────────────────────────────────────────
    if "bornes" in c:
        return ("U", max(10, round(cout / 500)))

    # ── Capteurs → U ──────────────────────────────────────
    if "capteur" in c:
        nums = re.findall(r'(\d+)', c)
        if nums:
            return ("U", int(nums[0]))
        return ("U", 1)

    # ── Siphons / dalots / bâche → ml ─────────────────────
    if any(kw in c for kw in ["siphon", "dalot", "bâche", "bach", "prise"]):
        qte = max(10, round(cout / 3000))
        return ("ml", qte)

    # ── Éclairage → U ─────────────────────────────────────
    if "éclairage" in c or "eclairage" in c:
        return ("U", 1)

    # ── Contrôleurs → U ───────────────────────────────────
    if "contrôleur" in c or "controleur" in c:
        nums = re.findall(r'(\d+)', c)
        if nums:
            return ("U", int(nums[0]))
        return ("U", 1)

    # ── Grilles → U ───────────────────────────────────────
    if "grille" in c:
        return ("U", 1)

    # ── Stabilisation berge → ml ──────────────────────────
    if "berge" in c or "stabilisation" in c:
        qte = max(20, round(cout / 4000))
        return ("ml", qte)

    # ── Expertise → U ─────────────────────────────────────
    if "expertise" in c:
        return ("U", 1)

    # ── Default → U (unité) ──────────────────────────────
    return ("U", 1)


# Compute indicators for all projects
project_indicators = []
for p in projects:
    unite, qte_prevue = infer_physical_indicator(
        p.get("consistance", ""),
        p.get("intitule_rubrique", ""),
        p.get("cout", 0)
    )
    project_indicators.append((unite, qte_prevue))


# ═══════════════════════════════════════════════════════════════
# COLUMN LAYOUT FOR WEEKLY SHEETS
# ═══════════════════════════════════════════════════════════════
# Section A: Identification (B-I) - locked reference
# Section B: Indicateurs Physiques (J-P) - input + auto
# Section C: Indicateurs Financiers (Q-X) - input + auto
# Section D: Écart + Obs (Y-Z)

WEEKLY_HEADERS = [
    # A: Identification
    "N°",                # B=2
    "Province",          # C=3
    "Commune",           # D=4
    "Rubrique",          # E=5
    "Intitulé Rubrique", # F=6
    "Projet",            # G=7
    "Consistance",       # H=8
    "Coût Total (DH)",   # I=9
    # B: Indicateurs Physiques
    "Unité",             # J=10
    "Qté Prévue",        # K=11
    "Qté Réalisée Sem.", # L=12  ← INPUT
    "Qté Réalisée Cumul",# M=13  ← AUTO
    "Avancement Physique %", # N=14  ← AUTO (M/K)
    "% Phys. Sem. Préc.", # O=15  ← AUTO (from prev sheet)
    "Accroissement Phys. %", # P=16  ← AUTO (N-O)
    # C: Indicateurs Financiers
    "Montant Décaissé (DH)", # Q=17  ← INPUT
    "Montant Payé (DH)",     # R=18  ← INPUT
    "Montant Ordonné (DH)",  # S=19  ← INPUT
    "Avancement Financier %", # T=20  ← INPUT
    "% Fin. Sem. Préc.",     # U=21  ← AUTO
    "Accroissement Fin. %",  # V=22  ← AUTO
    # D: Écart + Obs
    "Écart Phys.-Fin.",  # W=23  ← AUTO
    "Observations",      # X=24  ← INPUT
]

COL_START = 2
COL_END = COL_START + len(WEEKLY_HEADERS) - 1  # X = 24

# Column classifications
REF_COLS = list(range(2, 10))        # B-I: locked reference
REF_EDITABLE = [10, 11]              # J-K: Unité, Qté Prévue (editable in Liste Projets, locked in weekly)
PHYS_INPUT_COLS = [12]               # L: Qté Réalisée Sem.
PHYS_AUTO_COLS = [13, 14, 15, 16]    # M, N, O, P
FIN_INPUT_COLS = [17, 18, 19, 20]    # Q, R, S, T
FIN_AUTO_COLS = [21, 22]             # U, V
ECART_AUTO = [23]                    # W
OBS_INPUT = [24]                     # X

ALL_INPUT_COLS = PHYS_INPUT_COLS + FIN_INPUT_COLS + OBS_INPUT  # [12, 17, 18, 19, 20, 24]
ALL_AUTO_COLS = PHYS_AUTO_COLS + FIN_AUTO_COLS + ECART_AUTO    # [13,14,15,16,21,22,23]

WEEKLY_WIDTHS = {
    2: 5, 3: 14, 4: 18, 5: 9, 6: 32, 7: 28, 8: 34, 9: 14,
    10: 8, 11: 12, 12: 14, 13: 14, 14: 14, 15: 14, 16: 14,
    17: 16, 18: 16, 19: 16, 20: 14, 21: 14, 22: 14,
    23: 14, 24: 24,
}

DATA_ROW_START = 5
NUM_WEEKS = 4
WEEK_NAMES = [f"S{w:02d}" for w in range(1, NUM_WEEKS + 1)]

# ── Section header rows (row 3 is used for section group labels) ──
# We'll use a merged section header above the column headers


# ═══════════════════════════════════════════════════════════════
# Create Workbook
# ═══════════════════════════════════════════════════════════════
wb = Workbook()
wb.properties.creator = "Z.ai"


# ═══════════════════════════════════════════════════════════════
# SHEET 1: Liste Projets (Master Reference)
# ═══════════════════════════════════════════════════════════════
ws_ref = wb.active
ws_ref.title = "Liste Projets"
LAST_REF_COL = 11  # up to K

REF_HEADERS = [
    "N°", "Province", "Commune", "Rubrique", "Intitulé Rubrique",
    "Projet", "Consistance", "Coût Total (DH)",
    "Unité", "Qté Prévue"
]

setup_sheet(ws_ref, title="Liste Référentielle des Projets - Indicateurs Physiques", last_col=LAST_REF_COL)

# Section label row (row 3)
ws_ref.cell(row=3, column=2, value="Identification")
ws_ref.cell(row=3, column=2).font = Font(name=FONT_NAME, size=9, bold=True, color=NEUTRAL_600)
ws_ref.merge_cells(start_row=3, start_column=2, end_row=3, end_column=9)
ws_ref.cell(row=3, column=10, value="Indicateurs Physiques")
ws_ref.cell(row=3, column=10).font = Font(name=FONT_NAME, size=9, bold=True, color=ACCENT_POSITIVE)
ws_ref.merge_cells(start_row=3, start_column=10, end_row=3, end_column=11)

# Headers
for col_idx, h in enumerate(REF_HEADERS, start=2):
    ws_ref.cell(row=4, column=col_idx, value=h)
style_header_row(ws_ref, row_num=4, col_start=2, col_end=LAST_REF_COL)

for i, p in enumerate(projects):
    r = DATA_ROW_START + i
    unite, qte = project_indicators[i]

    ws_ref.cell(row=r, column=2, value=i + 1)
    ws_ref.cell(row=r, column=3, value=p["province"])
    ws_ref.cell(row=r, column=4, value=p["commune"])
    ws_ref.cell(row=r, column=5, value=p["rubrique"])
    ws_ref.cell(row=r, column=6, value=p["intitule_rubrique"])
    ws_ref.cell(row=r, column=7, value=p.get("intitule_projet", ""))
    ws_ref.cell(row=r, column=8, value=p["consistance"])
    ws_ref.cell(row=r, column=9, value=p["cout"])
    ws_ref.cell(row=r, column=9).number_format = '#,##0'
    ws_ref.cell(row=r, column=10, value=unite)
    ws_ref.cell(row=r, column=11, value=qte)
    ws_ref.cell(row=r, column=11).number_format = '#,##0'

    style_data_row(ws_ref, row_num=r, col_start=2, col_end=LAST_REF_COL, row_index=i)

    # Province color accent
    prov = p["province"]
    if prov in PROVINCE_FILLS:
        ws_ref.cell(row=r, column=3).fill = PROVINCE_FILLS[prov]
        ws_ref.cell(row=r, column=3).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PROVINCE_COLORS.get(prov, NEUTRAL_900))

    # Unité + Qté Prévue in green accent
    ws_ref.cell(row=r, column=10).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=ACCENT_POSITIVE)
    ws_ref.cell(row=r, column=11).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=ACCENT_POSITIVE)

# Totals
tr = DATA_ROW_START + N
ws_ref.cell(row=tr, column=2, value="TOTAL")
ws_ref.cell(row=tr, column=9).value = f"=SUM(I{DATA_ROW_START}:I{tr - 1})"
ws_ref.cell(row=tr, column=9).number_format = '#,##0'
style_total_row(ws_ref, row_num=tr, col_start=2, col_end=LAST_REF_COL)

REF_WIDTHS = {2: 5, 3: 14, 4: 18, 5: 9, 6: 32, 7: 28, 8: 34, 9: 14, 10: 8, 11: 12}
for c, w in REF_WIDTHS.items():
    ws_ref.column_dimensions[get_column_letter(c)].width = w

ws_ref.freeze_panes = "C5"
ws_ref.sheet_properties.tabColor = PRIMARY


# ═══════════════════════════════════════════════════════════════
# SHEETS S01-S04: Weekly Entry Sheets
# ═══════════════════════════════════════════════════════════════
for week_idx, week_name in enumerate(WEEK_NAMES):
    ws = wb.create_sheet(title=week_name)

    week_label = f"Semaine {week_idx + 1}"
    title_text = f"Suivi Hebdomadaire - {week_label} - Exécution Physique et Financière"
    setup_sheet(ws, title=title_text, last_col=COL_END)

    # Row 3: Section group labels + date input
    ws.row_dimensions[3].height = 22

    # Date input
    ws.cell(row=3, column=2, value="Date :")
    ws.cell(row=3, column=2).font = Font(name=FONT_NAME, size=10, bold=HEADER_BOLD, color=PRIMARY)
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=4)
    date_cell = ws.cell(row=3, column=3)
    date_cell.number_format = 'YYYY-MM-DD'
    date_cell.font = Font(name=FONT_NAME, size=11, color="0000FF")
    date_cell.fill = PatternFill("solid", fgColor="FFFFCC")
    date_cell.protection = Protection(locked=False)

    # Section headers (visual grouping)
    # Identification: B-I
    ws.merge_cells(start_row=3, start_column=5, end_row=3, end_column=9)
    ws.cell(row=3, column=5, value="IDENTIFICATION")
    ws.cell(row=3, column=5).font = Font(name=FONT_NAME, size=9, bold=True, color=NEUTRAL_600)
    ws.cell(row=3, column=5).alignment = Alignment(horizontal="center")

    # Indicateurs Physiques: J-P
    ws.merge_cells(start_row=3, start_column=10, end_row=3, end_column=16)
    ws.cell(row=3, column=10, value="INDICATEURS PHYSIQUES")
    ws.cell(row=3, column=10).font = Font(name=FONT_NAME, size=9, bold=True, color=ACCENT_POSITIVE)
    ws.cell(row=3, column=10).alignment = Alignment(horizontal="center")
    ws.cell(row=3, column=10).fill = PatternFill("solid", fgColor="E8F5E9")

    # Indicateurs Financiers: Q-V
    ws.merge_cells(start_row=3, start_column=17, end_row=3, end_column=22)
    ws.cell(row=3, column=17, value="INDICATEURS FINANCIERS")
    ws.cell(row=3, column=17).font = Font(name=FONT_NAME, size=9, bold=True, color=ACCENT_WARNING)
    ws.cell(row=3, column=17).alignment = Alignment(horizontal="center")
    ws.cell(row=3, column=17).fill = PatternFill("solid", fgColor="FEF9E7")

    # Écart + Obs: W-X
    ws.merge_cells(start_row=3, start_column=23, end_row=3, end_column=24)
    ws.cell(row=3, column=23, value="SYNTHÈSE")
    ws.cell(row=3, column=23).font = Font(name=FONT_NAME, size=9, bold=True, color=ACCENT_NEGATIVE)
    ws.cell(row=3, column=23).alignment = Alignment(horizontal="center")
    ws.cell(row=3, column=23).fill = PatternFill("solid", fgColor="FDEDEC")

    # Column headers (row 4)
    for col_idx, h in enumerate(WEEKLY_HEADERS, start=COL_START):
        ws.cell(row=4, column=col_idx, value=h)
    style_header_row(ws, row_num=4, col_start=COL_START, col_end=COL_END)

    # Sub-color the physical columns header in green tint
    for c in range(10, 17):
        cell = ws.cell(row=4, column=c)
        cell.fill = PatternFill("solid", fgColor="1B7D46")  # Darker green for physical header
        cell.font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color="FFFFFF")

    # Sub-color the financial columns header in amber tint
    for c in range(17, 23):
        cell = ws.cell(row=4, column=c)
        cell.fill = PatternFill("solid", fgColor="B8650A")  # Darker amber for financial header
        cell.font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color="FFFFFF")

    # Écart + Obs header
    for c in [23, 24]:
        cell = ws.cell(row=4, column=c)
        cell.fill = PatternFill("solid", fgColor="A0322B")  # Darker red
        cell.font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color="FFFFFF")

    # ── Data rows ──────────────────────────────────────────────
    for i, p in enumerate(projects):
        r = DATA_ROW_START + i
        unite, qte = project_indicators[i]

        # A: Identification (B-I) - locked reference
        ws.cell(row=r, column=2, value=i + 1)
        ws.cell(row=r, column=3, value=p["province"])
        ws.cell(row=r, column=4, value=p["commune"])
        ws.cell(row=r, column=5, value=p["rubrique"])
        ws.cell(row=r, column=6, value=p["intitule_rubrique"])
        ws.cell(row=r, column=7, value=p.get("intitule_projet", ""))
        ws.cell(row=r, column=8, value=p["consistance"])
        ws.cell(row=r, column=9, value=p["cout"])
        ws.cell(row=r, column=9).number_format = '#,##0'

        # B: Indicateurs Physiques
        # J=10: Unité (reference from Liste Projets)
        ws.cell(row=r, column=10, value=unite)
        ws.cell(row=r, column=10).font = Font(name=FONT_NAME, size=11, color=ACCENT_POSITIVE)

        # K=11: Qté Prévue (reference from Liste Projets)
        ws.cell(row=r, column=11, value=qte)
        ws.cell(row=r, column=11).number_format = '#,##0'
        ws.cell(row=r, column=11).font = Font(name=FONT_NAME, size=11, color=ACCENT_POSITIVE)

        # L=12: Qté Réalisée Semaine ← INPUT
        ws.cell(row=r, column=12).number_format = '#,##0.0'
        ws.cell(row=r, column=12).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=12).font = Font(name=FONT_NAME, size=11, color="0000FF")

        # M=13: Qté Réalisée Cumul ← AUTO
        if week_idx == 0:
            # First week: cumulative = this week's realization
            ws.cell(row=r, column=13).value = f"=L{r}"
        else:
            prev_sheet = WEEK_NAMES[week_idx - 1]
            ws.cell(row=r, column=13).value = f"='{prev_sheet}'!M{r}+L{r}"
            ws.cell(row=r, column=13).font = Font(name=FONT_NAME, size=11, color="008000")
        ws.cell(row=r, column=13).number_format = '#,##0.0'

        # N=14: Avancement Physique % ← AUTO = M/K
        ws.cell(row=r, column=14).value = f"=IFERROR(M{r}/K{r},0)"
        ws.cell(row=r, column=14).number_format = '0.0%'

        # O=15: % Physique Sem. Préc. ← AUTO
        if week_idx == 0:
            ws.cell(row=r, column=15, value=0)
        else:
            prev_sheet = WEEK_NAMES[week_idx - 1]
            ws.cell(row=r, column=15).value = f"='{prev_sheet}'!N{r}"
            ws.cell(row=r, column=15).font = Font(name=FONT_NAME, size=11, color="008000")
        ws.cell(row=r, column=15).number_format = '0.0%'

        # P=16: Accroissement Physique % ← AUTO = N - O
        ws.cell(row=r, column=16).value = f"=N{r}-O{r}"
        ws.cell(row=r, column=16).number_format = '0.0%'

        # C: Indicateurs Financiers
        # Q=17: Montant Décaissé ← INPUT
        ws.cell(row=r, column=17).number_format = '#,##0'
        ws.cell(row=r, column=17).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=17).font = Font(name=FONT_NAME, size=11, color="0000FF")

        # R=18: Montant Payé ← INPUT
        ws.cell(row=r, column=18).number_format = '#,##0'
        ws.cell(row=r, column=18).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=18).font = Font(name=FONT_NAME, size=11, color="0000FF")

        # S=19: Montant Ordonné ← INPUT
        ws.cell(row=r, column=19).number_format = '#,##0'
        ws.cell(row=r, column=19).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=19).font = Font(name=FONT_NAME, size=11, color="0000FF")

        # T=20: Avancement Financier % ← INPUT
        ws.cell(row=r, column=20).number_format = '0.0%'
        ws.cell(row=r, column=20).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=20).font = Font(name=FONT_NAME, size=11, color="0000FF")

        # U=21: % Financier Sem. Préc. ← AUTO
        if week_idx == 0:
            ws.cell(row=r, column=21, value=0)
        else:
            prev_sheet = WEEK_NAMES[week_idx - 1]
            ws.cell(row=r, column=21).value = f"='{prev_sheet}'!T{r}"
            ws.cell(row=r, column=21).font = Font(name=FONT_NAME, size=11, color="008000")
        ws.cell(row=r, column=21).number_format = '0.0%'

        # V=22: Accroissement Financier % ← AUTO
        ws.cell(row=r, column=22).value = f"=T{r}-U{r}"
        ws.cell(row=r, column=22).number_format = '0.0%'

        # D: Écart + Obs
        # W=23: Écart Phys.-Fin. ← AUTO
        ws.cell(row=r, column=23).value = f"=N{r}-T{r}"
        ws.cell(row=r, column=23).number_format = '0.0%'

        # X=24: Observations ← INPUT
        ws.cell(row=r, column=24).fill = PatternFill("solid", fgColor="FFFFCC")
        ws.cell(row=r, column=24).font = Font(name=FONT_NAME, size=11, color="0000FF")

        # ── Base styling ────────────────────────────────────────
        style_data_row(ws, row_num=r, col_start=COL_START, col_end=COL_END, row_index=i)

        # Province color accent on column C
        prov = p["province"]
        if prov in PROVINCE_FILLS:
            ws.cell(row=r, column=3).fill = PROVINCE_FILLS[prov]
            ws.cell(row=r, column=3).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PROVINCE_COLORS.get(prov, NEUTRAL_900))

        # Re-apply INPUT cell styling (overwritten by style_data_row)
        for ic in ALL_INPUT_COLS:
            c_cell = ws.cell(row=r, column=ic)
            c_cell.fill = PatternFill("solid", fgColor="FFFFCC")
            c_cell.font = Font(name=FONT_NAME, size=11, color="0000FF")

        # Re-apply AUTO cell styling
        for ac in ALL_AUTO_COLS:
            c_cell = ws.cell(row=r, column=ac)
            c_cell.font = font_body()

        # Cross-sheet ref styling for cumulative & sem. préc.
        if week_idx > 0:
            ws.cell(row=r, column=13).font = Font(name=FONT_NAME, size=11, color="008000")
            ws.cell(row=r, column=15).font = Font(name=FONT_NAME, size=11, color="008000")
            ws.cell(row=r, column=21).font = Font(name=FONT_NAME, size=11, color="008000")

        # Unité + Qté Prévue in green
        ws.cell(row=r, column=10).font = Font(name=FONT_NAME, size=11, color=ACCENT_POSITIVE)
        ws.cell(row=r, column=11).font = Font(name=FONT_NAME, size=11, color=ACCENT_POSITIVE)

        # Number formats (re-apply after style_data_row)
        ws.cell(row=r, column=9).number_format = '#,##0'
        ws.cell(row=r, column=11).number_format = '#,##0'
        ws.cell(row=r, column=12).number_format = '#,##0.0'
        ws.cell(row=r, column=13).number_format = '#,##0.0'
        for pct_col in [14, 15, 16, 20, 21, 22, 23]:
            ws.cell(row=r, column=pct_col).number_format = '0.0%'
        for amt_col in [17, 18, 19]:
            ws.cell(row=r, column=amt_col).number_format = '#,##0'

    # ── Totals row ─────────────────────────────────────────────
    total_r = DATA_ROW_START + N
    ws.cell(row=total_r, column=2, value="TOTAL")

    # Coût total
    ws.cell(row=total_r, column=9).value = f"=SUM(I{DATA_ROW_START}:I{total_r - 1})"
    ws.cell(row=total_r, column=9).number_format = '#,##0'

    # Qté Réalisée Semaine
    ws.cell(row=total_r, column=12).value = f"=SUM(L{DATA_ROW_START}:L{total_r - 1})"
    ws.cell(row=total_r, column=12).number_format = '#,##0.0'

    # Qté Réalisée Cumul
    ws.cell(row=total_r, column=13).value = f"=SUM(M{DATA_ROW_START}:M{total_r - 1})"
    ws.cell(row=total_r, column=13).number_format = '#,##0.0'

    # Avancement Physique % (weighted by cost)
    ws.cell(row=total_r, column=14).value = f'=IFERROR(SUMPRODUCT(N{DATA_ROW_START}:N{total_r-1},I{DATA_ROW_START}:I{total_r-1})/I{total_r},0)'
    ws.cell(row=total_r, column=14).number_format = '0.0%'

    # % Physique Sem. Préc. (weighted)
    ws.cell(row=total_r, column=15).value = f'=IFERROR(SUMPRODUCT(O{DATA_ROW_START}:O{total_r-1},I{DATA_ROW_START}:I{total_r-1})/I{total_r},0)'
    ws.cell(row=total_r, column=15).number_format = '0.0%'

    # Accroissement Physique %
    ws.cell(row=total_r, column=16).value = f"=N{total_r}-O{total_r}"
    ws.cell(row=total_r, column=16).number_format = '0.0%'

    # Montant Décaissé
    ws.cell(row=total_r, column=17).value = f"=SUM(Q{DATA_ROW_START}:Q{total_r - 1})"
    ws.cell(row=total_r, column=17).number_format = '#,##0'

    # Montant Payé
    ws.cell(row=total_r, column=18).value = f"=SUM(R{DATA_ROW_START}:R{total_r - 1})"
    ws.cell(row=total_r, column=18).number_format = '#,##0'

    # Montant Ordonné
    ws.cell(row=total_r, column=19).value = f"=SUM(S{DATA_ROW_START}:S{total_r - 1})"
    ws.cell(row=total_r, column=19).number_format = '#,##0'

    # Avancement Financier % (weighted)
    ws.cell(row=total_r, column=20).value = f'=IFERROR(SUMPRODUCT(T{DATA_ROW_START}:T{total_r-1},I{DATA_ROW_START}:I{total_r-1})/I{total_r},0)'
    ws.cell(row=total_r, column=20).number_format = '0.0%'

    # % Financier Sem. Préc. (weighted)
    ws.cell(row=total_r, column=21).value = f'=IFERROR(SUMPRODUCT(U{DATA_ROW_START}:U{total_r-1},I{DATA_ROW_START}:I{total_r-1})/I{total_r},0)'
    ws.cell(row=total_r, column=21).number_format = '0.0%'

    # Accroissement Financier %
    ws.cell(row=total_r, column=22).value = f"=T{total_r}-U{total_r}"
    ws.cell(row=total_r, column=22).number_format = '0.0%'

    # Écart
    ws.cell(row=total_r, column=23).value = f"=N{total_r}-T{total_r}"
    ws.cell(row=total_r, column=23).number_format = '0.0%'

    style_total_row(ws, row_num=total_r, col_start=COL_START, col_end=COL_END)

    # ── Conditional formatting ──────────────────────────────────
    # Accroissement Physique (P=16): green > 0, red < 0
    ws.conditional_formatting.add(
        f'P{DATA_ROW_START}:P{total_r - 1}',
        CellIsRule(operator='greaterThan', formula=['0'],
                   font=CF_POSITIVE_FONT, fill=CF_POSITIVE_FILL)
    )
    ws.conditional_formatting.add(
        f'P{DATA_ROW_START}:P{total_r - 1}',
        CellIsRule(operator='lessThan', formula=['0'],
                   font=CF_NEGATIVE_FONT, fill=CF_NEGATIVE_FILL)
    )

    # Accroissement Financier (V=22)
    ws.conditional_formatting.add(
        f'V{DATA_ROW_START}:V{total_r - 1}',
        CellIsRule(operator='greaterThan', formula=['0'],
                   font=CF_POSITIVE_FONT, fill=CF_POSITIVE_FILL)
    )
    ws.conditional_formatting.add(
        f'V{DATA_ROW_START}:V{total_r - 1}',
        CellIsRule(operator='lessThan', formula=['0'],
                   font=CF_NEGATIVE_FONT, fill=CF_NEGATIVE_FILL)
    )

    # Écart Phys.-Fin. (W=23)
    ws.conditional_formatting.add(
        f'W{DATA_ROW_START}:W{total_r - 1}',
        CellIsRule(operator='greaterThan', formula=['0.05'],
                   font=CF_POSITIVE_FONT, fill=CF_POSITIVE_FILL)
    )
    ws.conditional_formatting.add(
        f'W{DATA_ROW_START}:W{total_r - 1}',
        CellIsRule(operator='lessThan', formula=['-0.05'],
                   font=CF_NEGATIVE_FONT, fill=CF_NEGATIVE_FILL)
    )
    ws.conditional_formatting.add(
        f'W{DATA_ROW_START}:W{total_r - 1}',
        CellIsRule(operator='between', formula=['-0.05', '0.05'],
                   font=CF_WARNING_FONT, fill=CF_WARNING_FILL)
    )

    # Data bars on Avancement Physique % (N=14)
    ws.conditional_formatting.add(
        f'N{DATA_ROW_START}:N{total_r - 1}',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                    color=ACCENT_POSITIVE, showValue=True)
    )

    # Data bars on Avancement Financier % (T=20)
    ws.conditional_formatting.add(
        f'T{DATA_ROW_START}:T{total_r - 1}',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                    color=ACCENT_WARNING, showValue=True)
    )

    # ── Data validation ────────────────────────────────────────
    # % validation (0-100%)
    dv_pct = DataValidation(type="decimal", operator="between", formula1=0, formula2=1, allow_blank=True)
    dv_pct.error = "Valeur entre 0% et 100%"
    dv_pct.errorTitle = "Invalide"
    dv_pct.prompt = "Entrez 0.25 pour 25%"
    dv_pct.promptTitle = "Avancement %"
    ws.add_data_validation(dv_pct)
    dv_pct.add(f'T{DATA_ROW_START}:T{total_r - 1}')

    # Quantity validation (>=0)
    dv_qte = DataValidation(type="decimal", operator="greaterThanOrEqual", formula1=0, allow_blank=True)
    dv_qte.error = "Quantité doit être >= 0"
    dv_qte.errorTitle = "Invalide"
    dv_qte.prompt = "Entrez la quantité réalisée cette semaine"
    dv_qte.promptTitle = "Qté Réalisée"
    ws.add_data_validation(dv_qte)
    dv_qte.add(f'L{DATA_ROW_START}:L{total_r - 1}')

    # ── Column widths ──────────────────────────────────────────
    for c, w in WEEKLY_WIDTHS.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── Freeze panes ───────────────────────────────────────────
    ws.freeze_panes = "C5"

    # ── Protection ─────────────────────────────────────────────
    ws.protection.sheet = True
    ws.protection.password = 'ormvag2026'
    for i in range(N):
        r = DATA_ROW_START + i
        # Lock reference + auto cols
        for c in REF_COLS + REF_EDITABLE + ALL_AUTO_COLS:
            ws.cell(row=r, column=c).protection = Protection(locked=True)
        # Unlock input cols
        for c in ALL_INPUT_COLS:
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    ws.cell(row=3, column=3).protection = Protection(locked=False)

    # Tab colors
    tab_colors = ["4472C4", "ED7D31", "70AD47", "FFC000"]
    ws.sheet_properties.tabColor = tab_colors[week_idx % len(tab_colors)]


# ═══════════════════════════════════════════════════════════════
# SHEET: Récapitulatif Hebdomadaire
# ═══════════════════════════════════════════════════════════════
ws_recap = wb.create_sheet(title="Récapitulatif")
LAST_RECAP_COL = 14

setup_sheet(ws_recap, title="Récapitulatif Hebdomadaire - Suivi d'Exécution", last_col=LAST_RECAP_COL)

last_week = WEEK_NAMES[-1]
tr_last = DATA_ROW_START + N

# ── Section 1: KPI ────────────────────────────────────────────
kpi_labels = ["Indicateur", "Valeur"]
for col_idx, label in enumerate(kpi_labels, start=2):
    ws_recap.cell(row=4, column=col_idx, value=label)
style_header_row(ws_recap, row_num=4, col_start=2, col_end=3)

kpi_data = [
    ("Total Projets", N),
    ("Coût Total (DH)", f"='{last_week}'!I{tr_last}"),
    ("Avancement Physique Moy. %", f"='{last_week}'!N{tr_last}"),
    ("Avancement Financier Moy. %", f"='{last_week}'!T{tr_last}"),
    ("Montant Décaissé Total (DH)", f"='{last_week}'!Q{tr_last}"),
    ("Montant Payé Total (DH)", f"='{last_week}'!R{tr_last}"),
    ("Écart Physique-Financier", f"='{last_week}'!W{tr_last}"),
]

for ki, (label, val) in enumerate(kpi_data):
    r = 5 + ki
    ws_recap.cell(row=r, column=2, value=label)
    ws_recap.cell(row=r, column=2).font = font_body()
    ws_recap.cell(row=r, column=3, value=val)
    ws_recap.cell(row=r, column=3).font = font_kpi()
    ws_recap.cell(row=r, column=3).alignment = align_number()
    if "%" in label:
        ws_recap.cell(row=r, column=3).number_format = '0.0%'
    elif "DH" in label:
        ws_recap.cell(row=r, column=3).number_format = '#,##0'
    ws_recap.cell(row=r, column=3).font = Font(name=FONT_NAME, size=11, color="008000")

# ── Section 2: Évolution Hebdomadaire ─────────────────────────
trend_start = 14
ws_recap.cell(row=trend_start, column=2, value="Évolution Hebdomadaire")
ws_recap.cell(row=trend_start, column=2).font = font_subheader()
ws_recap.merge_cells(start_row=trend_start, start_column=2, end_row=trend_start, end_column=9)

trend_hdr_row = trend_start + 1
trend_headers = [
    "Semaine", "Avanc. Phys. %", "Accroiss. Phys. %",
    "Qté Réalisée Cumul", "Avanc. Fin. %", "Accroiss. Fin. %",
    "Décaissé (DH)", "Écart Phys.-Fin."
]

for col_idx, h in enumerate(trend_headers, start=2):
    ws_recap.cell(row=trend_hdr_row, column=col_idx, value=h)
style_header_row(ws_recap, row_num=trend_hdr_row, col_start=2, col_end=9)

for wi, wn in enumerate(WEEK_NAMES):
    r = trend_hdr_row + 1 + wi
    tr = DATA_ROW_START + N
    ws_recap.cell(row=r, column=2, value=wn)
    ws_recap.cell(row=r, column=3).value = f"='{wn}'!N{tr}"
    ws_recap.cell(row=r, column=3).number_format = '0.0%'
    ws_recap.cell(row=r, column=4).value = f"='{wn}'!P{tr}"
    ws_recap.cell(row=r, column=4).number_format = '0.0%'
    ws_recap.cell(row=r, column=5).value = f"='{wn}'!M{tr}"
    ws_recap.cell(row=r, column=5).number_format = '#,##0.0'
    ws_recap.cell(row=r, column=6).value = f"='{wn}'!T{tr}"
    ws_recap.cell(row=r, column=6).number_format = '0.0%'
    ws_recap.cell(row=r, column=7).value = f"='{wn}'!V{tr}"
    ws_recap.cell(row=r, column=7).number_format = '0.0%'
    ws_recap.cell(row=r, column=8).value = f"='{wn}'!Q{tr}"
    ws_recap.cell(row=r, column=8).number_format = '#,##0'
    ws_recap.cell(row=r, column=9).value = f"='{wn}'!W{tr}"
    ws_recap.cell(row=r, column=9).number_format = '0.0%'

    style_data_row(ws_recap, row_num=r, col_start=2, col_end=9, row_index=wi)
    for col in range(3, 10):
        ws_recap.cell(row=r, column=col).font = Font(name=FONT_NAME, size=11, color="008000")

# Conditional formatting on trend
td_start = trend_hdr_row + 1
td_end = trend_hdr_row + NUM_WEEKS
for trend_col in ['D', 'G']:
    ws_recap.conditional_formatting.add(
        f'{trend_col}{td_start}:{trend_col}{td_end}',
        CellIsRule(operator='greaterThan', formula=['0'],
                   font=CF_POSITIVE_FONT, fill=CF_POSITIVE_FILL)
    )
    ws_recap.conditional_formatting.add(
        f'{trend_col}{td_start}:{trend_col}{td_end}',
        CellIsRule(operator='lessThan', formula=['0'],
                   font=CF_NEGATIVE_FONT, fill=CF_NEGATIVE_FILL)
    )

# ── Section 3: Résumé par Province ────────────────────────────
prov_start = td_end + 3
ws_recap.cell(row=prov_start, column=2, value="Résumé par Province")
ws_recap.cell(row=prov_start, column=2).font = font_subheader()
ws_recap.merge_cells(start_row=prov_start, start_column=2, end_row=prov_start, end_column=9)

prov_hdr_row = prov_start + 1
prov_headers = [
    "Province", "Nb Projets", "Coût Total (DH)", "Avanc. Phys. %",
    "Qté Réalisée Cumul", "Avanc. Fin. %",
    "Décaissé (DH)", "Écart Phys.-Fin."
]

for col_idx, h in enumerate(prov_headers, start=2):
    ws_recap.cell(row=prov_hdr_row, column=col_idx, value=h)
style_header_row(ws_recap, row_num=prov_hdr_row, col_start=2, col_end=9)

provinces = ["Kénitra", "Sidi Kacem", "Sidi Slimane"]
for pi, prov in enumerate(provinces):
    r = prov_hdr_row + 1 + pi
    prov_indices = [i for i, p in enumerate(projects) if p["province"] == prov]

    ws_recap.cell(row=r, column=2, value=prov)
    ws_recap.cell(row=r, column=2).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PROVINCE_COLORS.get(prov, NEUTRAL_900))
    ws_recap.cell(row=r, column=2).fill = PROVINCE_FILLS.get(prov, PatternFill())
    ws_recap.cell(row=r, column=3, value=len(prov_indices))

    # Coût total
    ws_recap.cell(row=r, column=4).value = f'=SUMPRODUCT((\'{last_week}\'!C{DATA_ROW_START}:C{DATA_ROW_START+N-1}="{prov}")*\'{last_week}\'!I{DATA_ROW_START}:I{DATA_ROW_START+N-1})'
    ws_recap.cell(row=r, column=4).number_format = '#,##0'

    # Avancement Physique % (weighted)
    ws_recap.cell(row=r, column=5).value = f'=IFERROR(SUMPRODUCT((\'{last_week}\'!C{DATA_ROW_START}:C{DATA_ROW_START+N-1}="{prov}")*\'{last_week}\'!N{DATA_ROW_START}:N{DATA_ROW_START+N-1}*\'{last_week}\'!I{DATA_ROW_START}:I{DATA_ROW_START+N-1})/SUMPRODUCT((\'{last_week}\'!C{DATA_ROW_START}:C{DATA_ROW_START+N-1}="{prov}")*\'{last_week}\'!I{DATA_ROW_START}:I{DATA_ROW_START+N-1}),0)'
    ws_recap.cell(row=r, column=5).number_format = '0.0%'

    # Qté Réalisée Cumul
    ws_recap.cell(row=r, column=6).value = f'=SUMPRODUCT((\'{last_week}\'!C{DATA_ROW_START}:C{DATA_ROW_START+N-1}="{prov}")*\'{last_week}\'!M{DATA_ROW_START}:M{DATA_ROW_START+N-1})'
    ws_recap.cell(row=r, column=6).number_format = '#,##0.0'

    # Avancement Financier % (weighted)
    ws_recap.cell(row=r, column=7).value = f'=IFERROR(SUMPRODUCT((\'{last_week}\'!C{DATA_ROW_START}:C{DATA_ROW_START+N-1}="{prov}")*\'{last_week}\'!T{DATA_ROW_START}:T{DATA_ROW_START+N-1}*\'{last_week}\'!I{DATA_ROW_START}:I{DATA_ROW_START+N-1})/SUMPRODUCT((\'{last_week}\'!C{DATA_ROW_START}:C{DATA_ROW_START+N-1}="{prov}")*\'{last_week}\'!I{DATA_ROW_START}:I{DATA_ROW_START+N-1}),0)'
    ws_recap.cell(row=r, column=7).number_format = '0.0%'

    # Décaissé
    ws_recap.cell(row=r, column=8).value = f'=SUMPRODUCT((\'{last_week}\'!C{DATA_ROW_START}:C{DATA_ROW_START+N-1}="{prov}")*\'{last_week}\'!Q{DATA_ROW_START}:Q{DATA_ROW_START+N-1})'
    ws_recap.cell(row=r, column=8).number_format = '#,##0'

    # Écart
    ws_recap.cell(row=r, column=9).value = f"=E{r}-G{r}"
    ws_recap.cell(row=r, column=9).number_format = '0.0%'

    style_data_row(ws_recap, row_num=r, col_start=2, col_end=9, row_index=pi)
    ws_recap.cell(row=r, column=2).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PROVINCE_COLORS.get(prov, NEUTRAL_900))
    ws_recap.cell(row=r, column=2).fill = PROVINCE_FILLS.get(prov, PatternFill())
    for col in range(4, 10):
        ws_recap.cell(row=r, column=col).font = Font(name=FONT_NAME, size=11, color="008000")

# Province totals
prov_total_row = prov_hdr_row + 1 + len(provinces)
ws_recap.cell(row=prov_total_row, column=2, value="TOTAL")
ws_recap.cell(row=prov_total_row, column=3).value = f"=SUM(C{prov_hdr_row+1}:C{prov_total_row-1})"
ws_recap.cell(row=prov_total_row, column=4).value = f"=SUM(D{prov_hdr_row+1}:D{prov_total_row-1})"
ws_recap.cell(row=prov_total_row, column=4).number_format = '#,##0'
ws_recap.cell(row=prov_total_row, column=5).value = f"='{last_week}'!N{tr_last}"
ws_recap.cell(row=prov_total_row, column=5).number_format = '0.0%'
ws_recap.cell(row=prov_total_row, column=6).value = f"=SUM(F{prov_hdr_row+1}:F{prov_total_row-1})"
ws_recap.cell(row=prov_total_row, column=6).number_format = '#,##0.0'
ws_recap.cell(row=prov_total_row, column=7).value = f"='{last_week}'!T{tr_last}"
ws_recap.cell(row=prov_total_row, column=7).number_format = '0.0%'
ws_recap.cell(row=prov_total_row, column=8).value = f"=SUM(H{prov_hdr_row+1}:H{prov_total_row-1})"
ws_recap.cell(row=prov_total_row, column=8).number_format = '#,##0'
ws_recap.cell(row=prov_total_row, column=9).value = f"=E{prov_total_row}-G{prov_total_row}"
ws_recap.cell(row=prov_total_row, column=9).number_format = '0.0%'
style_total_row(ws_recap, row_num=prov_total_row, col_start=2, col_end=9)

# ── Section 4: Détail Unités Physiques par Province ──────────
units_start = prov_total_row + 3
ws_recap.cell(row=units_start, column=2, value="Détail des Unités Physiques Utilisées")
ws_recap.cell(row=units_start, column=2).font = font_subheader()
ws_recap.merge_cells(start_row=units_start, start_column=2, end_row=units_start, end_column=6)

units_hdr_row = units_start + 1
units_headers = ["Unité", "Nb Projets", "Qté Prévue Totale", "Qté Réalisée Cumul", "Avancement %"]
for col_idx, h in enumerate(units_headers, start=2):
    ws_recap.cell(row=units_hdr_row, column=col_idx, value=h)
style_header_row(ws_recap, row_num=units_hdr_row, col_start=2, col_end=6)

# Group projects by unit
from collections import defaultdict
unit_groups = defaultdict(list)
for i, (unite, qte) in enumerate(project_indicators):
    unit_groups[unite].append(i)

for ui, (unite, indices) in enumerate(sorted(unit_groups.items())):
    r = units_hdr_row + 1 + ui
    ws_recap.cell(row=r, column=2, value=unite)
    ws_recap.cell(row=r, column=2).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=ACCENT_POSITIVE)
    ws_recap.cell(row=r, column=3, value=len(indices))

    # Qté Prévue Totale
    qte_prevue_total = sum(project_indicators[idx][1] for idx in indices)
    ws_recap.cell(row=r, column=4, value=qte_prevue_total)
    ws_recap.cell(row=r, column=4).number_format = '#,##0'

    # Qté Réalisée Cumul (from last week sheet)
    # Build SUMPRODUCT formula matching unité in column J
    ws_recap.cell(row=r, column=5).value = f'=SUMPRODUCT((\'{last_week}\'!J{DATA_ROW_START}:J{DATA_ROW_START+N-1}="{unite}")*\'{last_week}\'!M{DATA_ROW_START}:M{DATA_ROW_START+N-1})'
    ws_recap.cell(row=r, column=5).number_format = '#,##0.0'

    # Avancement %
    ws_recap.cell(row=r, column=6).value = f"=IFERROR(E{r}/D{r},0)"
    ws_recap.cell(row=r, column=6).number_format = '0.0%'

    style_data_row(ws_recap, row_num=r, col_start=2, col_end=6, row_index=ui)
    ws_recap.cell(row=r, column=2).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=ACCENT_POSITIVE)
    ws_recap.cell(row=r, column=5).font = Font(name=FONT_NAME, size=11, color="008000")
    ws_recap.cell(row=r, column=6).font = Font(name=FONT_NAME, size=11, color="008000")

# Data bar on unit avancement
units_data_start = units_hdr_row + 1
units_data_end = units_hdr_row + len(unit_groups)
ws_recap.conditional_formatting.add(
    f'F{units_data_start}:F{units_data_end}',
    DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                color=ACCENT_POSITIVE, showValue=True)
)

RECAP_WIDTHS = {2: 24, 3: 12, 4: 18, 5: 18, 6: 18, 7: 18, 8: 18, 9: 16}
for c, w in RECAP_WIDTHS.items():
    ws_recap.column_dimensions[get_column_letter(c)].width = w

ws_recap.freeze_panes = "C5"
ws_recap.sheet_properties.tabColor = "7030A0"


# ═══════════════════════════════════════════════════════════════
# SHEET: Guide d'utilisation
# ═══════════════════════════════════════════════════════════════
ws_guide = wb.create_sheet(title="Guide")
setup_sheet(ws_guide, title="Guide d'Utilisation du Canevas Hebdomadaire", last_col=6)

guide_items = [
    ("CODE COULEUR", "Jaune = Saisie | Vert = Formule auto inter-feuille | Noir = Formule calculée | Bleu = Valeur saisie"),
    ("SAISIE HEBDOMADAIRE", "Dans les feuilles S01-S04, saisissez uniquement les cellules jaunes : Qté Réalisée Semaine, Montants, Avancement Financier %, Observations"),
    ("INDICATEURS PHYSIQUES", "Chaque projet a une Unité (ml, km, m², m³, U) et une Qté Prévue définies dans 'Liste Projets'. L'Avancement Physique % est calculé automatiquement : Qté Réalisée Cumul / Qté Prévue"),
    ("CUMUL AUTOMATIQUE", "La Qté Réalisée Cumul = Cumul sem. précédente + Qté réalisée cette semaine. Le lien entre semaines est automatique."),
    ("AJOUTER SEMAINES", "Dupliquez S04, renommez en S05. Modifiez les formules des colonnes M (Cumul), O (% Préc.), U (% Fin. Préc.) pour référencer S04."),
    ("PROTECTION", "Cellules de référence et formules verrouillées. Mot de passe : ormvag2026"),
    ("MODIFIER UNITÉS/QTÉ", "Allez dans 'Liste Projets' pour modifier les Unités et Qté Prévue. Ces valeurs sont reportées dans les feuilles hebdomadaires."),
    ("FORMAT POURCENTAGE", "Saisir 0.25 pour 25%, 0.50 pour 50%, 1.00 pour 100%. Valeurs entre 0 et 1."),
    ("ÉCART PHYS.-FIN.", "Positif = le physique avance plus vite que le financier. Négatif = le financier avance plus vite. Problème si > 5% d'écart."),
]

for gi, (rubrique, desc) in enumerate(guide_items):
    r = 4 + gi
    ws_guide.cell(row=r, column=2, value=rubrique)
    ws_guide.cell(row=r, column=2).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PRIMARY)
    ws_guide.cell(row=r, column=3, value=desc)
    ws_guide.cell(row=r, column=3).font = font_body()
    ws_guide.cell(row=r, column=3).alignment = Alignment(wrap_text=True, vertical="center")
    ws_guide.row_dimensions[r].height = 32

ws_guide.column_dimensions['B'].width = 24
ws_guide.column_dimensions['C'].width = 90
ws_guide.sheet_properties.tabColor = NEUTRAL_600


# ═══════════════════════════════════════════════════════════════
# Reorder sheets
# ═══════════════════════════════════════════════════════════════
desired_order = ["Liste Projets", "Guide"] + WEEK_NAMES + ["Récapitulatif"]
sheet_map = {ws.title: ws for ws in wb.worksheets}
wb._sheets = [sheet_map[name] for name in desired_order]


# ═══════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════
OUTPUT_PATH = "/home/z/my-project/download/Canevas_Hebdomadaire_ORMVAG_Gharb.xlsx"
wb.save(OUTPUT_PATH)
print(f"Workbook saved to: {OUTPUT_PATH}")
print(f"Projects: {N}")
print(f"Physical units used: {dict((u, len(idxs)) for u, idxs in sorted(unit_groups.items()))}")
print(f"Weekly sheets: {', '.join(WEEK_NAMES)}")
print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
