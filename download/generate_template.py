import sys, os
sys.path.insert(0, os.path.join(os.path.dirname(__file__), '..', 'skills', 'xlsx', 'templates'))
from base import *

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule, DataBarRule

use_palette_explicit("bloomberg")

wb = Workbook()

# ===================================================================
# FEUILLE 1: Suivi des Projets
# ===================================================================
ws1 = wb.active
ws1.title = "Suivi des Projets"

headers1 = [
    "N°",
    "Province",
    "Commune",
    "Rubrique",
    "Intitulé Rubrique",
    "Intitulé Projet",
    "Consistance",
    "Budget Prévu (DH)",
    "Avancement Physique (%)",
    "Avancement Financier (%)",
    "Montant Décaissé (DH)",
    "Reste à Décaisser (DH)",
    "Taux de Décaissement (%)",
    "Écart Phys.-Fin. (pts)",
    "Statut",
    "Observations",
]

last_col1 = len(headers1) + 1  # +1 because starting from column B

setup_sheet(ws1, title="Suivi Physique et Financier des Projets — Région du Gharb 2026", last_col=last_col1)

# Write headers
for col_idx, header in enumerate(headers1, start=2):
    ws1.cell(row=4, column=col_idx, value=header)
style_header_row(ws1, row_num=4, col_start=2, col_end=last_col1)

# Data rows - 72 projects
projects_data = [
    ["Sidi Kacem","KHENICHET","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage des collecteurs d'assainissement du secteur EST2",2091568],
    ["Sidi Kacem","KHENICHET","40-21","Travaux de pistes","Travaux de réhabilitation des pistes agricoles","Entretien des pistes agricoles Douar Nouagra et Douar ouled Salem",78000],
    ["Sidi Kacem","KHENICHET","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Curage des prises sur l'oued, bâche d'aspiration et bassin de régulation.","Curage de la chambre des joints de démontage des siphonnes de la station de pompage EST2.",126000],
    ["Sidi Kacem","KHENICHET","50-21","Entretien et réparation des stations de pompage","Prestation de maintenance des débitmètres et leurs équipements annexes","Changement trois débitmètre Ø 600mm inondés de la station de pompage EST 2 et deux capteurs VEGASON 62",165500],
    ["Sidi Kacem","BIR TALEB","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage du canal d'assainisement Tihli Aval",1000000],
    ["Sidi Kacem","BIR TALEB","50-23","Entretien et réparation du réseau d'assainissement et de drainage","","Curage des collecteurs d'assainissement du secteur P11/1",846731],
    ["Sidi Kacem","BIR TALEB","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Curage des prises sur l'oued, bâche d'aspiration et bassin de régulation.","Curage des prises sur l'oued de la station de pompage P11/1.",200000],
    ["Sidi Kacem","BIR TALEB","50-21","Entretien et réparation des stations de pompage","Prestation de maintenance des débitmètres et leurs équipements annexes","Changement trois débitmètre Ø 400mm inondés de la station de pompage P11/1",141900],
    ["Sidi Kacem","BIR TALEB","50-21","Entretien et réparation des stations de pompage","Prestations de maintenance et de réparation mécanique des pompes et des groupes motopompes submersibles","Démontage et expertise des roues des pompes principaux et pompes a vide de la station P11/1",26700],
    ["Sidi Kacem","BIR TALEB","50-21","Entretien et réparation des stations de pompage","Stabilisation Berge oued","Pose de gabions et recouvrement du talus par matelas Reno des stations de pompage P11/1.",480000],
    ["Sidi Kacem","BIR TALEB","50-21","Entretien et réparation des stations de pompage","Génie civile des locaux des stations de pompage.","Expertise d'état et d'étanchéité des locaux pompes.",300000],
    ["Sidi Kacem","CHEBANAT","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage des collecteurs d'assainissement des secteur SH5 et SH6.",1000000],
    ["Sidi Kacem","CHEBANAT","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux révetés","Curage du canal principal d'irrigation beht partie aval 1.",739200],
    ["Sidi Kacem","ZIRARA","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux révetés","Curage du canal principal d'irrigation beht partie aval 2.",1276800],
    ["Sidi Kacem","DAR LAASLOUJI","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage des collecteurs d'assainissement des secteurs P8-1 et P8-2",5796360],
    ["Sidi Kacem","DAR LAASLOUJI","40-21","Travaux de pistes","Entretien des pistes agricoles","Entretien des pistes agricoles",340000],
    ["Sidi Kacem","HOUAFATE","20-11","Travaux de réhabilitation des équipements","Travaux de réhabilitation des canaux portés d'irrigation","Fourniture et de pose de canaux porté et accessoires du réseau S3 du secteur S13",397934],
    ["Sidi Kacem","HOUAFATE","20-11","Travaux de réhabilitation des équipements","","Fourniture et de pose de canaux porté et accessoires du réseau T2S1 du secteur S13",207575],
    ["Sidi Kacem","HOUAFATE","40-21","Travaux de pistes","Entretien des pistes agricoles","Entretien des pistes agricoles",2200000],
    ["Sidi Kacem","HOUAFATE","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Curage des prises sur l'oued, bâche d'aspiration et bassin de régulation.","Curage des prises sur l'oued de la station de pompage S17 et S13.",300000],
    ["Sidi Kacem","HOUAFATE","50-21","Entretien et réparation des stations de pompage","Prestation de maintenance des débitmètres et leurs équipements annexes","Changement trois débitmètre Ø 600mm inondés de la station de pompage S13.",250000],
    ["Sidi Kacem","HOUAFATE","50-21","Entretien et réparation des stations de pompage","Prestations de maintenance des moteurs électriques des stations de pompage.","Révision des trois moteurs électriques 200 KW principaux de la station S13.",25000],
    ["Sidi Kacem","HOUAFATE","50-21","Entretien et réparation des stations de pompage","Prestations de maintenance et de réparation mécanique des pompes et des groupes motopompes submersibles","Démontage et expertise des roues des pompes principaux et pompes a vide de la station S13.",38000],
    ["Sidi Kacem","HOUAFATE","50-21","Entretien et réparation des stations de pompage","Stabilisation Berge oued","Pose de gabions et recouvrement du talus par matelas Reno des stations de pompage S13.",10560000],
    ["Sidi Kacem","HOUAFATE","50-21","Entretien et réparation des stations de pompage","Mur clôture des stations de pompage.","Stabilisation du sol et reconstruction des murs clôture (270m) cote oued de la station S17.",144000],
    ["Sidi Kacem","HOUAFATE","50-21","Entretien et réparation des stations de pompage","Génie civile des locaux des stations de pompage.","Expertise d'état et d'étanchéité des locaux pompes, commande, puissance des stations de pompage S17 et S13.",800000],
    ["Sidi Kacem","NOUIRAT","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage des collecteurs d'assainissement des secteurs EST1 et N5",1733306],
    ["Sidi Kacem","NOUIRAT","40-21","Travaux de pistes","Entretien des pistes agricoles","Entretien des pistes agricoles",670000],
    ["Sidi Kacem","NOUIRAT","20-11","Travaux de réhabilitation des équipements","Travaux de réhabilitation des ouvrages de génie civil","Réhabilitation de l'ouvrage de l'ouvrage de débouche SE1",2500000],
    ["Sidi Kacem","NOUIRAT","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Curage des prises sur l'oued, bâche d'aspiration et bassin de régulation.","Curage des prises sur l'oued de la station de pompage E1.",800000],
    ["Sidi Kacem","NOUIRAT","50-21","Entretien et réparation des stations de pompage","Prestation de maintenance des débitmètres et leurs équipements annexes","Changement du débitmètre Ø 1200mm inondé de la station de pompage E1.",195000],
    ["Sidi Kacem","NOUIRAT","50-21","Entretien et réparation des stations de pompage","Prestations de maintenance des moteurs électriques des stations de pompage.","Révision des deux moteurs électriques principaux de la station E1.",50000],
    ["Sidi Kacem","NOUIRAT","50-21","Entretien et réparation des stations de pompage","Prestations de maintenance et de réparation mécanique des pompes et des groupes motopompes submersibles","Démontage et expertise des cinq pompes principaux et pompes a vide de la station E1.",245000],
    ["Sidi Kacem","NOUIRAT","50-21","Entretien et réparation des stations de pompage","Génie civile des locaux des stations de pompage.","Expertise d'état et d'étanchéité des locaux pompes de la station de pompage E1.",1000000],
    ["Sidi Kacem","RMILAT","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage des collecteurs d'assainissement des secteurs P7",5024032],
    ["Sidi Kacem","RMILAT","40-21","Travaux de pistes","Entretien des pistes agricoles","Entretien des pistes agricoles",400000],
    ["Sidi Kacem","SAFSAF","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage des collecteurs d'assainissement des secteurs S7, S9 et S11",6000000],
    ["Sidi Kacem","SAFSAF","20-11","Travaux de réhabilitation des équipements","Travaux de réhabilitation des canaux portés d'irrigation","Fourniture et pose de canaux porté et accessoires du réseau S4 du secteur S11",232266],
    ["Sidi Kacem","SAFSAF","20-11","Travaux de réhabilitation des équipements","","Recalage des canaux portés du réseau T5S5 du secteur S9",1055025],
    ["Sidi Kacem","SAFSAF","20-11","Travaux de réhabilitation des équipements","","Recalage des canaux portés du réseau T7S5 du secteur S9",1004535],
    ["Sidi Kacem","SAFSAF","20-11","Travaux de réhabilitation des équipements","","Recalage des canaux portés du réseau T2S5 du secteur S9",793830],
    ["Sidi Kacem","SAFSAF","40-21","Travaux de pistes","Entretien des pistes agricoles","Entretien des pistes agricoles",2000000],
    ["Sidi Kacem","SAFSAF","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Curage des prises sur l'oued, bâche d'aspiration et bassin de régulation.","Curage des prises sur l'oued des stations de pompage S11, S9, 7D et 7B.",200000],
    ["Sidi Kacem","SAFSAF","50-21","Entretien et réparation des stations de pompage","Prestation de maintenance des débitmètres et leurs équipements annexes","Changements Sept des débitmètres Ø 600mm inondé des deux stations de pompage S11 et S9.",580000],
    ["Sidi Kacem","SAFSAF","50-21","Entretien et réparation des stations de pompage","Stabilisation Berge oued","Pose de murs en gabions et recouvrement du talus par matelas Reno des stations de pompage S11, S9, 7D et 7B.",4000000],
    ["Sidi Kacem","SIDI AZZOUZ","40-21","Travaux de pistes","Entretien des pistes agricoles","Entretien des pistes agricoles",655000],
    ["Sidi Kacem","SIDI EL KAMEL","20-11","Travaux de réhabilitation des équipements","Travaux de réhabilitation des canaux portés d'irrigation","Fourniture et de pose de canaux porté et accessoires du réseau SR4 du secteur S3R",5182920],
    ["Sidi Kacem","SIDI EL KAMEL","20-11","Travaux de réhabilitation des équipements","","Fourniture et de pose de canaux porté et accessoires du réseau SR6 du secteur S1RIZ",152818],
    ["Sidi Kacem","SIDI EL KAMEL","20-11","Travaux de réhabilitation des équipements","","Recalage des canaux portés du réseau SR4 du secteur S3R",1418000],
    ["Sidi Kacem","SIDI EL KAMEL","20-11","Travaux de réhabilitation des équipements","Travaux de réhabilitation des ouvrages de génie civil","Réhabilitation des murs de soutènements des ouvrages de franchissements du canal d'assainissement CS3",800000],
    ["Sidi Kacem","SIDI EL KAMEL","40-21","Travaux de pistes","Entretien des pistes agricoles","Entretien des pistes agricoles",3700000],
    ["Sidi Kacem","SIDI EL KAMEL","50-21","Entretien et réparation des stations de pompage","Prestation de maintenance des débitmètres et leurs équipements annexes","Changement de quatre débitmètres Ø 500mm et trois Ø 800mm inondé des stations de pompage 1R, 3R et 5R.",705000],
    ["Sidi Kacem","SIDI EL KAMEL","50-21","Entretien et réparation des stations de pompage","Mur clôture des stations de pompage.","Stabilisation et reconstruction des murs clôture cote oued des stations de pompage 7A, 3R et 1R.",372000],
    ["Kénitra","Bni malek","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage des emissaires et colatures secondaire au niveau des secteurs MDA ET N9",10000000],
    ["Kénitra","Bni malek","40-21","Travaux de pistes","Travaux de réhabilitation des pistes agricoles","Entretien des pistes agricoles des secteurs MDA et N3",2000000],
    ["Kénitra","Bni malek","50-21","Entretien et réparation des stations de pompage","Prestations de maintenance et de réparation mécanique des pompes et des groupes motopompes submersibles","Station de pompage MDA secteur MDA",300000],
    ["Kénitra","Bni malek","20-11","Travaux de réhabilitation des équipements","Travaux de réhabilitation des ouvrages de génie civil","Rétablissement des bornes affaissées des secteurs N2 et N3",100000],
    ["Kénitra","SOUK TLET","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage des emissaires et colatures secondaire au niveau des secteurs N2,N3 ET N4",3000000],
    ["Kénitra","SOUK TLET","40-21","Travaux de pistes","Travaux de réhabilitation des pistes agricoles","Entretien des pistes des secteurs N2,N3,N4 et C4R",5000000],
    ["Kénitra","Sidi Med Lahmar","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage des emissaires et colatures secondaires secteurs N1",3000000],
    ["Kénitra","Sidi Med Lahmar","40-21","Travaux de pistes","Travaux de réhabilitation des pistes agricoles","Entretien des pistes agricoles des secteurs N9 et N1",3000000],
    ["Kénitra","Mograne","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage des canaux d'assainissement Hors secteur équipé ghoufera et laison beht sebou",6000000],
    ["Kénitra","Mograne","40-21","Travaux de pistes","Travaux de réhabilitation des pistes agricoles","Entretien des pistes agricoles des secteurs SEBOU1,SEBOU2, BEHT3 et BEHT4",6000000],
    ["Kénitra","Sidi allal tazi","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage du canal d'assainissement CK",3000000],
    ["Kénitra","AMEUR SAFLIA","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage des canaux d'assainissement amont Tiflt Sebou",1000000],
    ["Kénitra","Oulad SLAMA","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage des canaux d'assainissement debouche Tiflt aval",1000000],
    ["Kénitra","BHARA Oulad AYAD","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage des canaux d'assainissement canal nador bas et haut segment",6600000],
    ["Sidi Slimane","OULED HCEINE","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage des collecteurs d'assainissement du secteur EST4.",1805475],
    ["Sidi Slimane","OULED HCEINE","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des canaux en terre","Curage du canal d'assainissement Tihli central.",2587200],
    ["Sidi Slimane","OULED HCEINE","40-21","Travaux de pistes","Travaux de réhabilitation des pistes agricoles","Entretien des pistes agricoles",3500000],
    ["Sidi Slimane","OULED HCEINE","50-21","Entretien et réparation des stations de pompage","Travaux de réparation des équipements hydromécaniques et électriques et entretien de la partie génie civil","Démontage et expertise des roues des pompes de la station EST4 et P11/2",870000],
    ["Sidi Slimane","OULED HCEINE","50-23","Entretien et réparation du réseau d'assainissement et de drainage","Travaux de curage des prises des stations de pompage.","Curage des siphones des station de pompage EST2 et P11/2.",500000],
]

for i, proj in enumerate(projects_data):
    row_num = 5 + i
    # N°
    ws1.cell(row=row_num, column=2, value=i + 1)
    # Province, Commune, Rubrique, Intitulé Rubrique, Intitulé Projet, Consistance
    for j, val in enumerate(proj):
        ws1.cell(row=row_num, column=3 + j, value=val)
    # Budget Prévu (DH) - column I (9)
    ws1.cell(row=row_num, column=9, value=proj[6])
    ws1.cell(row=row_num, column=9).number_format = '#,##0'
    # Avancement Physique (%) - column J (10) - EMPTY for user input
    # Avancement Financier (%) - column K (11) - EMPTY for user input
    # Montant Décaissé (DH) - column L (12) - EMPTY for user input
    # Reste à Décaisser = Budget - Décaissé  (formula)
    budget_col = 'I'
    decaisse_col = 'L'
    ws1.cell(row=row_num, column=13).value = f'=IF({decaisse_col}{row_num}="","",{budget_col}{row_num}-{decaisse_col}{row_num})'
    ws1.cell(row=row_num, column=13).number_format = '#,##0'
    # Taux de Décaissement (%) = Décaissé / Budget (formula)
    ws1.cell(row=row_num, column=14).value = f'=IFERROR(IF({decaisse_col}{row_num}="","",{decaisse_col}{row_num}/{budget_col}{row_num}),"")'
    ws1.cell(row=row_num, column=14).number_format = '0.0%'
    # Écart Phys.-Fin. = Avancement Physique - Avancement Financier (formula)
    phys_col = 'J'
    fin_col = 'K'
    ws1.cell(row=row_num, column=15).value = f'=IFERROR(IF(OR({phys_col}{row_num}="",{fin_col}{row_num}=""),"",{phys_col}{row_num}-{fin_col}{row_num}),"")'
    ws1.cell(row=row_num, column=15).number_format = '0.0'
    # Statut - column P (16) - EMPTY for user input (with dropdown)
    # Observations - column Q (17) - EMPTY

    style_data_row(ws1, row_num=row_num, col_start=2, col_end=last_col1, row_index=i)

# Number format for percentage columns
for i in range(len(projects_data)):
    row_num = 5 + i
    ws1.cell(row=row_num, column=10).number_format = '0.0%'
    ws1.cell(row=row_num, column=11).number_format = '0.0%'

# Totals row
total_row = 5 + len(projects_data)
ws1.cell(row=total_row, column=2, value="TOTAL")
ws1.cell(row=total_row, column=9).value = f'=SUM(I5:I{total_row-1})'
ws1.cell(row=total_row, column=9).number_format = '#,##0'
ws1.cell(row=total_row, column=12).value = f'=SUM(L5:L{total_row-1})'
ws1.cell(row=total_row, column=12).number_format = '#,##0'
ws1.cell(row=total_row, column=13).value = f'=IFERROR(L{total_row}/I{total_row},"")'
ws1.cell(row=total_row, column=13).number_format = '#,##0'
ws1.cell(row=total_row, column=14).value = f'=IFERROR(L{total_row}/I{total_row},"")'
ws1.cell(row=total_row, column=14).number_format = '0.0%'
# Average avancement
ws1.cell(row=total_row, column=10).value = f'=IFERROR(AVERAGE(J5:J{total_row-1}),"")'
ws1.cell(row=total_row, column=10).number_format = '0.0%'
ws1.cell(row=total_row, column=11).value = f'=IFERROR(AVERAGE(K5:K{total_row-1}),"")'
ws1.cell(row=total_row, column=11).number_format = '0.0%'
ws1.cell(row=total_row, column=15).value = f'=IFERROR(AVERAGE(O5:O{total_row-1}),"")'
ws1.cell(row=total_row, column=15).number_format = '0.0'
style_total_row(ws1, row_num=total_row, col_start=2, col_end=last_col1)

# Data validation: Province dropdown
dv_province = DataValidation(type="list", formula1='"Kénitra,Sidi Kacem,Sidi Slimane"', allow_blank=True)
dv_province.prompt = "Sélectionner la province"
dv_province.promptTitle = "Province"
ws1.add_data_validation(dv_province)
dv_province.add(f'C5:C{total_row-1}')

# Data validation: Statut dropdown
dv_statut = DataValidation(type="list", formula1='"Terminé,En cours,Non démarré"', allow_blank=True)
dv_statut.prompt = "Sélectionner le statut du projet"
dv_statut.promptTitle = "Statut"
ws1.add_data_validation(dv_statut)
dv_statut.add(f'P5:P{total_row-1}')

# Data validation: Avancement physique (0-100%)
dv_pct = DataValidation(type="decimal", operator="between", formula1="0", formula2="1", allow_blank=True)
dv_pct.prompt = "Saisir le pourcentage d'avancement (0 à 100%)"
dv_pct.promptTitle = "Avancement"
ws1.add_data_validation(dv_pct)
dv_pct.add(f'J5:J{total_row-1}')
dv_pct.add(f'K5:K{total_row-1}')

# Conditional formatting: Statut
green_fill = PatternFill(bgColor="E8F5E9")
green_font = Font(color="1B7D46")
red_fill = PatternFill(bgColor="FDEDEC")
red_font = Font(color="C0392B")
amber_fill = PatternFill(bgColor="FEF9E7")
amber_font = Font(color="D4820A")

ws1.conditional_formatting.add(f'P5:P{total_row-1}',
    CellIsRule(operator='equal', formula=['"Terminé"'], fill=green_fill, font=green_font))
ws1.conditional_formatting.add(f'P5:P{total_row-1}',
    CellIsRule(operator='equal', formula=['"En cours"'], fill=amber_fill, font=amber_font))
ws1.conditional_formatting.add(f'P5:P{total_row-1}',
    CellIsRule(operator='equal', formula=['"Non démarré"'], fill=red_fill, font=red_font))

# Conditional formatting: Avancement color scale
ws1.conditional_formatting.add(f'J5:J{total_row-1}',
    ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                   mid_type='num', mid_value=0.5, mid_color='FFEB84',
                   end_type='num', end_value=1, end_color='63BE7B'))
ws1.conditional_formatting.add(f'K5:K{total_row-1}',
    ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                   mid_type='num', mid_value=0.5, mid_color='FFEB84',
                   end_type='num', end_value=1, end_color='63BE7B'))

# Conditional formatting: Écart > 15pts → red
ws1.conditional_formatting.add(f'O5:O{total_row-1}',
    CellIsRule(operator='lessThan', formula=['-15'], fill=red_fill, font=red_font))

# Data bars on Budget
ws1.conditional_formatting.add(f'I5:I{total_row-1}',
    DataBarRule(start_type='min', end_type='max', color=PRIMARY, showValue=True))

# Column widths
ws1.column_dimensions['A'].width = 3
ws1.column_dimensions['B'].width = 5
ws1.column_dimensions['C'].width = 16
ws1.column_dimensions['D'].width = 18
ws1.column_dimensions['E'].width = 10
ws1.column_dimensions['F'].width = 32
ws1.column_dimensions['G'].width = 32
ws1.column_dimensions['H'].width = 40
ws1.column_dimensions['I'].width = 18
ws1.column_dimensions['J'].width = 16
ws1.column_dimensions['K'].width = 16
ws1.column_dimensions['L'].width = 18
ws1.column_dimensions['M'].width = 18
ws1.column_dimensions['N'].width = 16
ws1.column_dimensions['O'].width = 14
ws1.column_dimensions['P'].width = 14
ws1.column_dimensions['Q'].width = 30

# Freeze panes
ws1.freeze_panes = 'C5'

# ===================================================================
# FEUILLE 2: Récapitulatif par Province
# ===================================================================
ws2 = wb.create_sheet("Récap. par Province")

headers2 = [
    "Province",
    "Nb Projets",
    "Nb Communes",
    "Budget Total (DH)",
    "Montant Décaissé (DH)",
    "Reste à Décaisser (DH)",
    "Taux Décaissement (%)",
    "Avancement Physique Moyen (%)",
    "Avancement Financier Moyen (%)",
    "Écart Phys.-Fin. (pts)",
]

last_col2 = len(headers2) + 1
setup_sheet(ws2, title="Récapitulatif par Province — Région du Gharb 2026", last_col=last_col2)

for col_idx, header in enumerate(headers2, start=2):
    ws2.cell(row=4, column=col_idx, value=header)
style_header_row(ws2, row_num=4, col_start=2, col_end=last_col2)

provinces = ["Kénitra", "Sidi Kacem", "Sidi Slimane"]
for i, prov in enumerate(provinces):
    row_num = 5 + i
    ws2.cell(row=row_num, column=2, value=prov)
    # Formulas referencing Suivi des Projets sheet
    ws2.cell(row=row_num, column=3).value = f'=COUNTIF(\'Suivi des Projets\'!C5:C76,B{row_num})'
    ws2.cell(row=row_num, column=4).value = f'=COUNTIFS(\'Suivi des Projets\'!C5:C76,B{row_num},\'Suivi des Projets\'!D5:D76,"<>")'
    ws2.cell(row=row_num, column=5).value = f'=SUMIF(\'Suivi des Projets\'!C5:C76,B{row_num},\'Suivi des Projets\'!I5:I76)'
    ws2.cell(row=row_num, column=5).number_format = '#,##0'
    ws2.cell(row=row_num, column=6).value = f'=SUMIF(\'Suivi des Projets\'!C5:C76,B{row_num},\'Suivi des Projets\'!L5:L76)'
    ws2.cell(row=row_num, column=6).number_format = '#,##0'
    ws2.cell(row=row_num, column=7).value = f'=F{row_num}-E{row_num}'
    ws2.cell(row=row_num, column=7).number_format = '#,##0'
    ws2.cell(row=row_num, column=8).value = f'=IFERROR(F{row_num}/E{row_num},"")'
    ws2.cell(row=row_num, column=8).number_format = '0.0%'
    ws2.cell(row=row_num, column=9).value = f'=IFERROR(AVERAGEIF(\'Suivi des Projets\'!C5:C76,B{row_num},\'Suivi des Projets\'!J5:J76),"")'
    ws2.cell(row=row_num, column=9).number_format = '0.0%'
    ws2.cell(row=row_num, column=10).value = f'=IFERROR(AVERAGEIF(\'Suivi des Projets\'!C5:C76,B{row_num},\'Suivi des Projets\'!K5:K76),"")'
    ws2.cell(row=row_num, column=10).number_format = '0.0%'
    ws2.cell(row=row_num, column=11).value = f'=I{row_num}-J{row_num}'
    ws2.cell(row=row_num, column=11).number_format = '0.0'
    style_data_row(ws2, row_num=row_num, col_start=2, col_end=last_col2, row_index=i)

# Total row
total_row2 = 8
ws2.cell(row=total_row2, column=2, value="TOTAL RÉGION")
ws2.cell(row=total_row2, column=3).value = f'=SUM(C5:C7)'
ws2.cell(row=total_row2, column=5).value = f'=SUM(E5:E7)'
ws2.cell(row=total_row2, column=5).number_format = '#,##0'
ws2.cell(row=total_row2, column=6).value = f'=SUM(F5:F7)'
ws2.cell(row=total_row2, column=6).number_format = '#,##0'
ws2.cell(row=total_row2, column=7).value = f'=SUM(G5:G7)'
ws2.cell(row=total_row2, column=7).number_format = '#,##0'
ws2.cell(row=total_row2, column=8).value = f'=IFERROR(F{total_row2}/E{total_row2},"")'
ws2.cell(row=total_row2, column=8).number_format = '0.0%'
ws2.cell(row=total_row2, column=9).value = f'=IFERROR(AVERAGE(I5:I7),"")'
ws2.cell(row=total_row2, column=9).number_format = '0.0%'
ws2.cell(row=total_row2, column=10).value = f'=IFERROR(AVERAGE(J5:J7),"")'
ws2.cell(row=total_row2, column=10).number_format = '0.0%'
ws2.cell(row=total_row2, column=11).value = f'=I{total_row2}-J{total_row2}'
ws2.cell(row=total_row2, column=11).number_format = '0.0'
style_total_row(ws2, row_num=total_row2, col_start=2, col_end=last_col2)

# Conditional formatting on province avancement
ws2.conditional_formatting.add('I5:I7',
    ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                   mid_type='num', mid_value=0.5, mid_color='FFEB84',
                   end_type='num', end_value=1, end_color='63BE7B'))
ws2.conditional_formatting.add('J5:J7',
    ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                   mid_type='num', mid_value=0.5, mid_color='FFEB84',
                   end_type='num', end_value=1, end_color='63BE7B'))

# Column widths
ws2.column_dimensions['A'].width = 3
ws2.column_dimensions['B'].width = 18
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 14
ws2.column_dimensions['E'].width = 18
ws2.column_dimensions['F'].width = 20
ws2.column_dimensions['G'].width = 20
ws2.column_dimensions['H'].width = 18
ws2.column_dimensions['I'].width = 22
ws2.column_dimensions['J'].width = 22
ws2.column_dimensions['K'].width = 16

ws2.freeze_panes = 'C5'

# ===================================================================
# FEUILLE 3: Récapitulatif par Secteur
# ===================================================================
ws3 = wb.create_sheet("Récap. par Secteur")

headers3 = [
    "Secteur",
    "Nb Projets",
    "Nb Communes",
    "Budget Total (DH)",
    "Montant Décaissé (DH)",
    "Reste à Décaisser (DH)",
    "Taux Décaissement (%)",
    "Avancement Physique Moyen (%)",
    "Avancement Financier Moyen (%)",
    "Écart Phys.-Fin. (pts)",
]

last_col3 = len(headers3) + 1
setup_sheet(ws3, title="Récapitulatif par Secteur — Région du Gharb 2026", last_col=last_col3)

for col_idx, header in enumerate(headers3, start=2):
    ws3.cell(row=4, column=col_idx, value=header)
style_header_row(ws3, row_num=4, col_start=2, col_end=last_col3)

secteurs = [
    "Assainissement & Drainage",
    "Pistes agricoles",
    "Stations de pompage",
    "Réhabilitation équipements",
]
secteur_full = [
    "Entretien et réparation du réseau d'assainissement et de drainage",
    "Travaux de pistes",
    "Entretien et réparation des stations de pompage",
    "Travaux de réhabilitation des équipements",
]

for i, (short, full) in enumerate(zip(secteurs, secteur_full)):
    row_num = 5 + i
    ws3.cell(row=row_num, column=2, value=short)
    ws3.cell(row=row_num, column=3).value = f'=COUNTIF(\'Suivi des Projets\'!F5:F76,"{full}")'
    ws3.cell(row=row_num, column=5).value = f'=SUMIF(\'Suivi des Projets\'!F5:F76,"{full}",\'Suivi des Projets\'!I5:I76)'
    ws3.cell(row=row_num, column=5).number_format = '#,##0'
    ws3.cell(row=row_num, column=6).value = f'=SUMIF(\'Suivi des Projets\'!F5:F76,"{full}",\'Suivi des Projets\'!L5:L76)'
    ws3.cell(row=row_num, column=6).number_format = '#,##0'
    ws3.cell(row=row_num, column=7).value = f'=F{row_num}-E{row_num}'
    ws3.cell(row=row_num, column=7).number_format = '#,##0'
    ws3.cell(row=row_num, column=8).value = f'=IFERROR(F{row_num}/E{row_num},"")'
    ws3.cell(row=row_num, column=8).number_format = '0.0%'
    ws3.cell(row=row_num, column=9).value = f'=IFERROR(AVERAGEIF(\'Suivi des Projets\'!F5:F76,"{full}",\'Suivi des Projets\'!J5:J76),"")'
    ws3.cell(row=row_num, column=9).number_format = '0.0%'
    ws3.cell(row=row_num, column=10).value = f'=IFERROR(AVERAGEIF(\'Suivi des Projets\'!F5:F76,"{full}",\'Suivi des Projets\'!K5:K76),"")'
    ws3.cell(row=row_num, column=10).number_format = '0.0%'
    ws3.cell(row=row_num, column=11).value = f'=I{row_num}-J{row_num}'
    ws3.cell(row=row_num, column=11).number_format = '0.0'
    style_data_row(ws3, row_num=row_num, col_start=2, col_end=last_col3, row_index=i)

# Total row
total_row3 = 9
ws3.cell(row=total_row3, column=2, value="TOTAL")
ws3.cell(row=total_row3, column=3).value = f'=SUM(C5:C8)'
ws3.cell(row=total_row3, column=5).value = f'=SUM(E5:E8)'
ws3.cell(row=total_row3, column=5).number_format = '#,##0'
ws3.cell(row=total_row3, column=6).value = f'=SUM(F5:F8)'
ws3.cell(row=total_row3, column=6).number_format = '#,##0'
ws3.cell(row=total_row3, column=7).value = f'=SUM(G5:G8)'
ws3.cell(row=total_row3, column=7).number_format = '#,##0'
ws3.cell(row=total_row3, column=8).value = f'=IFERROR(F{total_row3}/E{total_row3},"")'
ws3.cell(row=total_row3, column=8).number_format = '0.0%'
ws3.cell(row=total_row3, column=9).value = f'=IFERROR(AVERAGE(I5:I8),"")'
ws3.cell(row=total_row3, column=9).number_format = '0.0%'
ws3.cell(row=total_row3, column=10).value = f'=IFERROR(AVERAGE(J5:J8),"")'
ws3.cell(row=total_row3, column=10).number_format = '0.0%'
ws3.cell(row=total_row3, column=11).value = f'=I{total_row3}-J{total_row3}'
ws3.cell(row=total_row3, column=11).number_format = '0.0'
style_total_row(ws3, row_num=total_row3, col_start=2, col_end=last_col3)

# Conditional formatting
ws3.conditional_formatting.add('I5:I8',
    ColorScaleRule(start_type='num', start_value=0, start_color='F8696B',
                   mid_type='num', mid_value=0.5, mid_color='FFEB84',
                   end_type='num', end_value=1, end_color='63BE7B'))

ws3.column_dimensions['A'].width = 3
ws3.column_dimensions['B'].width = 28
ws3.column_dimensions['C'].width = 12
ws3.column_dimensions['D'].width = 14
ws3.column_dimensions['E'].width = 18
ws3.column_dimensions['F'].width = 20
ws3.column_dimensions['G'].width = 20
ws3.column_dimensions['H'].width = 18
ws3.column_dimensions['I'].width = 22
ws3.column_dimensions['J'].width = 22
ws3.column_dimensions['K'].width = 16

ws3.freeze_panes = 'C5'

# ===================================================================
# FEUILLE 4: Mode d'emploi
# ===================================================================
ws4 = wb.create_sheet("Mode d'emploi")

last_col4 = 8
setup_sheet(ws4, title="Mode d'emploi — Fichier de Suivi", last_col=last_col4)

instructions = [
    ["", "Ce fichier Excel sert de modèle pour alimenter le dashboard SIG Gharb avec les données de suivi physique et financier des projets."],
    ["", ""],
    ["COLONNES À REMPLIR", ""],
    ["Avancement Physique (%)", "Saisir le pourcentage de progression des travaux sur le terrain (0% à 100%). Exemples : 0%, 25%, 50%, 75%, 100%"],
    ["Avancement Financier (%)", "Saisir le pourcentage de consommation du budget (0% à 100%). Représente la part du budget engagée/dépensée."],
    ["Montant Décaissé (DH)", "Saisir le montant effectivement décaissé en Dirhams. Doit être inférieur ou égal au Budget Prévu."],
    ["Statut", "Sélectionner dans la liste déroulante : Terminé, En cours, ou Non démarré"],
    ["Observations", "Zone libre pour notes, commentaires, problèmes rencontrés, etc."],
    ["", ""],
    ["COLONNES CALCULÉES AUTOMATIQUEMENT", ""],
    ["Reste à Décaisser", "= Budget Prévu - Montant Décaissé"],
    ["Taux de Décaissement", "= Montant Décaissé / Budget Prévu"],
    ["Écart Phys.-Fin. (pts)", "= Avancement Physique - Avancement Financier. Un écart négatif important (>15pts) signale un retard physique."],
    ["", ""],
    ["COULEURS CONDITIONNELLES", ""],
    ["Avancement", "Rouge (0-33%) → Jaune (34-66%) → Vert (67-100%)"],
    ["Statut", "Vert = Terminé, Orange = En cours, Rouge = Non démarré"],
    ["Écart", "Rouge si écart < -15pts (retard physique significatif)"],
    ["", ""],
    ["FEUILLES DU FICHIER", ""],
    ["Suivi des Projets", "Liste détaillée des 72 projets avec toutes les informations"],
    ["Récap. par Province", "Synthèse automatique par province (Kénitra, Sidi Kacem, Sidi Slimane)"],
    ["Récap. par Secteur", "Synthèse automatique par secteur d'activité"],
    ["Mode d'emploi", "Cette feuille — guide d'utilisation"],
    ["", ""],
    ["IMPORT DANS LE DASHBOARD", ""],
    ["", "Pour importer les données dans le dashboard SIG Gharb :"],
    ["1.", "Remplir les colonnes Avancement Physique, Avancement Financier, Montant Décaissé et Statut pour chaque projet"],
    ["2.", "Exporter le fichier en format CSV ou JSON"],
    ["3.", "Mettre à jour le fichier dashboard_data.json avec les nouvelles valeurs"],
    ["4.", "Le dashboard se met à jour automatiquement après déploiement"],
]

for i, (label, desc) in enumerate(instructions):
    row_num = 5 + i
    ws4.cell(row=row_num, column=2, value=label)
    ws4.cell(row=row_num, column=3, value=desc)
    if label and label.isupper():
        ws4.cell(row=row_num, column=2).font = font_subheader()
    elif label and label[0].isdigit():
        ws4.cell(row=row_num, column=2).font = Font(name=FONT_NAME, size=11, bold=True, color=PRIMARY)
    else:
        ws4.cell(row=row_num, column=2).font = font_body()
    ws4.cell(row=row_num, column=3).font = font_body()
    ws4.cell(row=row_num, column=3).alignment = Alignment(wrap_text=True, vertical='center')

ws4.column_dimensions['A'].width = 3
ws4.column_dimensions['B'].width = 28
ws4.column_dimensions['C'].width = 80

# Save
output_path = "/home/z/my-project/download/Modele_Suivi_Avancement_Gharb_2026.xlsx"
wb.properties.creator = "Z.ai"
wb.save(output_path)
print(f"Saved to {output_path}")
