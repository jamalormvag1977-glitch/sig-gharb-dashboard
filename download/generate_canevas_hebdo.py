#!/usr/bin/env python3
"""
Canevas Hebdomadaire - Suivi d'Exécution Physique et Financière
ORMVAG - Région du Gharb
Génère un classeur Excel avec feuilles hebdomadaires liées automatiquement.
"""

import json
import sys
import os

# ── Skill imports ──────────────────────────────────────────────
XLSX_SKILL_DIR = "/home/z/my-project/skills/xlsx"
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment, numbers, Protection
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import CellIsRule, DataBarRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from copy import copy
from templates.base import (
    setup_sheet, style_header_row, style_data_row, style_total_row,
    font_title, font_header, font_subheader, font_body, font_caption, font_kpi, font_kpi_label,
    fill_header, fill_total, fill_data_row,
    border_header, border_total,
    align_title, align_header, align_number, align_text, align_date,
    COLUMN_WIDTHS, FORMATS, ROW_HEIGHTS,
    PRIMARY, PRIMARY_LIGHT, SECONDARY,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING,
    NEUTRAL_900, NEUTRAL_600, NEUTRAL_200, NEUTRAL_100, NEUTRAL_0,
    HEADER_BOLD, FONT_NAME, HEADER_TEXT,
    CF_POSITIVE_FILL, CF_POSITIVE_FONT, CF_NEGATIVE_FILL, CF_NEGATIVE_FONT,
    CF_WARNING_FILL, CF_WARNING_FONT,
    auto_fit_columns,
)

# Province colors from the GIS dashboard
PROVINCE_COLORS = {
    "Kénitra": "D4A017",       # Yellow gold
    "Sidi Kacem": "C76E7E",    # Pink
    "Sidi Slimane": "5BB58A",  # Mint green
}

# Province fills for conditional formatting
PROVINCE_FILLS = {
    "Kénitra": PatternFill("solid", fgColor="FFF8DC"),       # Light cornsilk
    "Sidi Kacem": PatternFill("solid", fgColor="FFF0F3"),    # Light pink
    "Sidi Slimane": PatternFill("solid", fgColor="F0FFF4"),  # Light honeydew
}

# ── Load project data ──────────────────────────────────────────
DATA_PATH = "/tmp/sig-gharb-extract/src/data/dashboard_data.json"
with open(DATA_PATH, "r", encoding="utf-8") as f:
    data = json.load(f)

projects = data["projects"]
N = len(projects)  # 72

# ── Column layout for weekly sheets ───────────────────────────
# B=N°, C=Province, D=Commune, E=Rubrique, F=Intitulé Rubrique,
# G=Projet, H=Consistance, I=Coût Total,
# J=Avancement Physique %, K=% Sem. Préc. (auto), L=Accroissement Physique % (auto),
# M=Montant Décaissé, N=Montant Payé, O=Montant Ordonné,
# P=Avancement Financier %, Q=% Sem. Préc. (auto), R=Accroissement Financier % (auto),
# S=Écart Physique-Financier (auto), T=Observations
WEEKLY_HEADERS = [
    "N°", "Province", "Commune", "Rubrique", "Intitulé Rubrique",
    "Projet", "Consistance", "Coût Total (DH)",
    "Avancement Physique %", "% Physique Sem. Préc.", "Accroissement Physique %",
    "Montant Décaissé (DH)", "Montant Payé (DH)", "Montant Ordonné (DH)",
    "Avancement Financier %", "% Financier Sem. Préc.", "Accroissement Financier %",
    "Écart Physique-Financier", "Observations"
]
COL_START = 2  # Column B
COL_END = COL_START + len(WEEKLY_HEADERS) - 1  # Column T = 20

# Column widths for weekly sheets
WEEKLY_WIDTHS = {
    2: 5,    # N°
    3: 16,   # Province
    4: 20,   # Commune
    5: 10,   # Rubrique
    6: 36,   # Intitulé Rubrique
    7: 32,   # Projet
    8: 36,   # Consistance
    9: 16,   # Coût Total
    10: 14,  # Avancement Physique %
    11: 14,  # % Physique Sem. Préc.
    12: 14,  # Accroissement Physique %
    13: 18,  # Montant Décaissé
    14: 18,  # Montant Payé
    15: 18,  # Montant Ordonné
    16: 14,  # Avancement Financier %
    17: 14,  # % Financier Sem. Préc.
    18: 14,  # Accroissement Financier %
    19: 14,  # Écart Physique-Financier
    20: 28,  # Observations
}

# Input columns (user fills in) - J, M, N, O, P, T
INPUT_COLS = [10, 13, 14, 15, 16, 20]  # J=Av.Phys%, M=Décaissé, N=Payé, O=Ordonné, P=Av.Fin%, T=Obs

# Auto-calculated columns - K, L, Q, R, S
AUTO_COLS = [11, 12, 17, 18, 19]

# Reference columns (locked, from Liste Projets) - B through I
REF_COLS = list(range(2, 10))

# Number of weekly sheets to create
NUM_WEEKS = 4
WEEK_NAMES = [f"S{w:02d}" for w in range(1, NUM_WEEKS + 1)]

# ── Data row start ─────────────────────────────────────────────
DATA_ROW_START = 5  # Row 5 is first data row (Row 1=margin, 2=title, 3=spacer, 4=header)

# ── Create Workbook ────────────────────────────────────────────
wb = Workbook()
wb.properties.creator = "Z.ai"

# ═══════════════════════════════════════════════════════════════
# SHEET 1: Liste Projets (Master Reference)
# ═══════════════════════════════════════════════════════════════
ws_ref = wb.active
ws_ref.title = "Liste Projets"
setup_sheet(ws_ref, title="Liste Référentielle des Projets - ORMVAG Gharb", last_col=9)

REF_HEADERS = ["N°", "Province", "Commune", "Rubrique", "Intitulé Rubrique", "Projet", "Consistance", "Coût Total (DH)"]

for col_idx, h in enumerate(REF_HEADERS, start=2):
    ws_ref.cell(row=4, column=col_idx, value=h)
style_header_row(ws_ref, row_num=4, col_start=2, col_end=9)

for i, p in enumerate(projects):
    r = DATA_ROW_START + i
    ws_ref.cell(row=r, column=2, value=i + 1)
    ws_ref.cell(row=r, column=3, value=p["province"])
    ws_ref.cell(row=r, column=4, value=p["commune"])
    ws_ref.cell(row=r, column=5, value=p["rubrique"])
    ws_ref.cell(row=r, column=6, value=p["intitule_rubrique"])
    ws_ref.cell(row=r, column=7, value=p.get("intitule_projet", ""))
    ws_ref.cell(row=r, column=8, value=p["consistance"])
    ws_ref.cell(row=r, column=9, value=p["cout"])
    ws_ref.cell(row=r, column=9).number_format = '#,##0'

    style_data_row(ws_ref, row_num=r, col_start=2, col_end=9, row_index=i)

    # Province color accent on column C
    prov = p["province"]
    if prov in PROVINCE_FILLS:
        ws_ref.cell(row=r, column=3).fill = PROVINCE_FILLS[prov]
        ws_ref.cell(row=r, column=3).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PROVINCE_COLORS.get(prov, NEUTRAL_900))

# Totals row
total_row = DATA_ROW_START + N
ws_ref.cell(row=total_row, column=2, value="TOTAL")
ws_ref.cell(row=total_row, column=9).value = f"=SUM(I{DATA_ROW_START}:I{total_row - 1})"
ws_ref.cell(row=total_row, column=9).number_format = '#,##0'
style_total_row(ws_ref, row_num=total_row, col_start=2, col_end=9)

# Column widths
REF_WIDTHS = {2: 5, 3: 16, 4: 20, 5: 10, 6: 36, 7: 32, 8: 36, 9: 16}
for c, w in REF_WIDTHS.items():
    ws_ref.column_dimensions[get_column_letter(c)].width = w

ws_ref.freeze_panes = "C5"
ws_ref.sheet_properties.tabColor = PRIMARY


# ═══════════════════════════════════════════════════════════════
# SHEETS 2-5: S01 to S04 (Weekly Entry Sheets)
# ═══════════════════════════════════════════════════════════════
for week_idx, week_name in enumerate(WEEK_NAMES):
    ws = wb.create_sheet(title=week_name)

    # Title
    week_label = f"Semaine {week_idx + 1}"
    title_text = f"Suivi Hebdomadaire - {week_label} - Exécution Physique et Financière"
    setup_sheet(ws, title=title_text, last_col=COL_END)

    # Add week date input row (Row 3)
    ws.cell(row=3, column=2, value="Date Semaine :")
    ws.cell(row=3, column=2).font = font_subheader()
    ws.cell(row=3, column=2).alignment = align_text()
    ws.merge_cells(start_row=3, start_column=3, end_row=3, end_column=4)
    date_cell = ws.cell(row=3, column=3)
    date_cell.number_format = 'YYYY-MM-DD'
    date_cell.font = Font(name=FONT_NAME, size=11, color="0000FF")  # Blue = manual input
    date_cell.alignment = align_date()
    date_cell.fill = PatternFill("solid", fgColor="FFFFCC")  # Light yellow = input cell

    # Headers
    for col_idx, h in enumerate(WEEKLY_HEADERS, start=COL_START):
        ws.cell(row=4, column=col_idx, value=h)
    style_header_row(ws, row_num=4, col_start=COL_START, col_end=COL_END)

    # Sub-header color band for column groups
    # Physical section: J-K-L (columns 10-12) → light green header accent
    # Financial section: M-N-O-P-Q-R (columns 13-18) → light amber header accent
    # Écart: S (column 19) → light red accent

    for i, p in enumerate(projects):
        r = DATA_ROW_START + i

        # Reference data (columns B-I) - locked, from Liste Projets
        ws.cell(row=r, column=2, value=i + 1)     # N°
        ws.cell(row=r, column=3, value=p["province"])  # Province
        ws.cell(row=r, column=4, value=p["commune"])   # Commune
        ws.cell(row=r, column=5, value=p["rubrique"])   # Rubrique
        ws.cell(row=r, column=6, value=p["intitule_rubrique"])  # Intitulé
        ws.cell(row=r, column=7, value=p.get("intitule_projet", ""))  # Projet
        ws.cell(row=r, column=8, value=p["consistance"])  # Consistance
        ws.cell(row=r, column=9, value=p["cout"])  # Coût
        ws.cell(row=r, column=9).number_format = '#,##0'

        # INPUT: Avancement Physique % (column J=10)
        cell_phys = ws.cell(row=r, column=10)
        cell_phys.number_format = '0.0%'
        cell_phys.fill = PatternFill("solid", fgColor="FFFFCC")  # Yellow = input
        cell_phys.font = Font(name=FONT_NAME, size=11, color="0000FF")  # Blue = input

        # AUTO: % Physique Sem. Préc. (column K=11)
        cell_phys_prev = ws.cell(row=r, column=11)
        cell_phys_prev.number_format = '0.0%'
        if week_idx == 0:
            # First week: 0% as baseline
            cell_phys_prev.value = 0
        else:
            prev_sheet = WEEK_NAMES[week_idx - 1]
            cell_phys_prev.value = f"='{prev_sheet}'!J{r}"
            cell_phys_prev.font = Font(name=FONT_NAME, size=11, color="008000")  # Green = cross-sheet ref

        # AUTO: Accroissement Physique % (column L=12)
        cell_phys_delta = ws.cell(row=r, column=12)
        cell_phys_delta.value = f"=J{r}-K{r}"
        cell_phys_delta.number_format = '0.0%'

        # INPUT: Montant Décaissé (column M=13)
        cell_decaisse = ws.cell(row=r, column=13)
        cell_decaisse.number_format = '#,##0'
        cell_decaisse.fill = PatternFill("solid", fgColor="FFFFCC")
        cell_decaisse.font = Font(name=FONT_NAME, size=11, color="0000FF")

        # INPUT: Montant Payé (column N=14)
        cell_paye = ws.cell(row=r, column=14)
        cell_paye.number_format = '#,##0'
        cell_paye.fill = PatternFill("solid", fgColor="FFFFCC")
        cell_paye.font = Font(name=FONT_NAME, size=11, color="0000FF")

        # INPUT: Montant Ordonné (column O=15)
        cell_ordonne = ws.cell(row=r, column=15)
        cell_ordonne.number_format = '#,##0'
        cell_ordonne.fill = PatternFill("solid", fgColor="FFFFCC")
        cell_ordonne.font = Font(name=FONT_NAME, size=11, color="0000FF")

        # INPUT: Avancement Financier % (column P=16)
        cell_fin = ws.cell(row=r, column=16)
        cell_fin.number_format = '0.0%'
        cell_fin.fill = PatternFill("solid", fgColor="FFFFCC")
        cell_fin.font = Font(name=FONT_NAME, size=11, color="0000FF")

        # AUTO: % Financier Sem. Préc. (column Q=17)
        cell_fin_prev = ws.cell(row=r, column=17)
        cell_fin_prev.number_format = '0.0%'
        if week_idx == 0:
            cell_fin_prev.value = 0
        else:
            prev_sheet = WEEK_NAMES[week_idx - 1]
            cell_fin_prev.value = f"='{prev_sheet}'!P{r}"
            cell_fin_prev.font = Font(name=FONT_NAME, size=11, color="008000")

        # AUTO: Accroissement Financier % (column R=18)
        cell_fin_delta = ws.cell(row=r, column=18)
        cell_fin_delta.value = f"=P{r}-Q{r}"
        cell_fin_delta.number_format = '0.0%'

        # AUTO: Écart Physique-Financier (column S=19)
        cell_ecart = ws.cell(row=r, column=19)
        cell_ecart.value = f"=J{r}-P{r}"
        cell_ecart.number_format = '0.0%'

        # INPUT: Observations (column T=20)
        cell_obs = ws.cell(row=r, column=20)
        cell_obs.fill = PatternFill("solid", fgColor="FFFFCC")
        cell_obs.font = Font(name=FONT_NAME, size=11, color="0000FF")

        # Apply data row styling (base)
        style_data_row(ws, row_num=r, col_start=COL_START, col_end=COL_END, row_index=i)

        # Province color accent on column C
        prov = p["province"]
        if prov in PROVINCE_FILLS:
            ws.cell(row=r, column=3).fill = PROVINCE_FILLS[prov]
            ws.cell(row=r, column=3).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PROVINCE_COLORS.get(prov, NEUTRAL_900))

        # Re-apply input cell styling after style_data_row overwrites it
        for ic in INPUT_COLS:
            c = ws.cell(row=r, column=ic)
            c.fill = PatternFill("solid", fgColor="FFFFCC")
            c.font = Font(name=FONT_NAME, size=11, color="0000FF")

        # Re-apply auto-cell styling
        for ac in AUTO_COLS:
            c = ws.cell(row=r, column=ac)
            if ac in [11, 17]:  # Cross-sheet references
                c.font = Font(name=FONT_NAME, size=11, color="008000") if week_idx > 0 else font_body()
            else:
                c.font = font_body()  # Calculated values = black

        # Number formatting
        ws.cell(row=r, column=9).number_format = '#,##0'
        for ac in [10, 11, 12, 16, 17, 18, 19]:
            ws.cell(row=r, column=ac).number_format = '0.0%'
        for ac in [13, 14, 15]:
            ws.cell(row=r, column=ac).number_format = '#,##0'

    # ── Totals row ──────────────────────────────────────────────
    tr = DATA_ROW_START + N
    ws.cell(row=tr, column=2, value="TOTAL")
    # Coût total
    ws.cell(row=tr, column=9).value = f"=SUM(I{DATA_ROW_START}:I{tr - 1})"
    ws.cell(row=tr, column=9).number_format = '#,##0'
    # Avancement Physique % moyen (weighted by cost)
    ws.cell(row=tr, column=10).value = f'=IFERROR(SUMPRODUCT(J{DATA_ROW_START}:J{tr-1},I{DATA_ROW_START}:I{tr-1})/I{tr},0)'
    ws.cell(row=tr, column=10).number_format = '0.0%'
    # % Physique Sem. Préc.
    ws.cell(row=tr, column=11).value = f'=IFERROR(SUMPRODUCT(K{DATA_ROW_START}:K{tr-1},I{DATA_ROW_START}:I{tr-1})/I{tr},0)'
    ws.cell(row=tr, column=11).number_format = '0.0%'
    # Accroissement Physique %
    ws.cell(row=tr, column=12).value = f"=J{tr}-K{tr}"
    ws.cell(row=tr, column=12).number_format = '0.0%'
    # Montant Décaissé
    ws.cell(row=tr, column=13).value = f"=SUM(M{DATA_ROW_START}:M{tr - 1})"
    ws.cell(row=tr, column=13).number_format = '#,##0'
    # Montant Payé
    ws.cell(row=tr, column=14).value = f"=SUM(N{DATA_ROW_START}:N{tr - 1})"
    ws.cell(row=tr, column=14).number_format = '#,##0'
    # Montant Ordonné
    ws.cell(row=tr, column=15).value = f"=SUM(O{DATA_ROW_START}:O{tr - 1})"
    ws.cell(row=tr, column=15).number_format = '#,##0'
    # Avancement Financier % moyen (weighted)
    ws.cell(row=tr, column=16).value = f'=IFERROR(SUMPRODUCT(P{DATA_ROW_START}:P{tr-1},I{DATA_ROW_START}:I{tr-1})/I{tr},0)'
    ws.cell(row=tr, column=16).number_format = '0.0%'
    # % Financier Sem. Préc.
    ws.cell(row=tr, column=17).value = f'=IFERROR(SUMPRODUCT(Q{DATA_ROW_START}:Q{tr-1},I{DATA_ROW_START}:I{tr-1})/I{tr},0)'
    ws.cell(row=tr, column=17).number_format = '0.0%'
    # Accroissement Financier %
    ws.cell(row=tr, column=18).value = f"=P{tr}-Q{tr}"
    ws.cell(row=tr, column=18).number_format = '0.0%'
    # Écart Physique-Financier
    ws.cell(row=tr, column=19).value = f"=J{tr}-P{tr}"
    ws.cell(row=tr, column=19).number_format = '0.0%'

    style_total_row(ws, row_num=tr, col_start=COL_START, col_end=COL_END)

    # ── Conditional formatting ──────────────────────────────────
    # Accroissement Physique (L): green if > 0, red if < 0
    ws.conditional_formatting.add(
        f'L{DATA_ROW_START}:L{tr - 1}',
        CellIsRule(operator='greaterThan', formula=['0'],
                   font=CF_POSITIVE_FONT, fill=CF_POSITIVE_FILL)
    )
    ws.conditional_formatting.add(
        f'L{DATA_ROW_START}:L{tr - 1}',
        CellIsRule(operator='lessThan', formula=['0'],
                   font=CF_NEGATIVE_FONT, fill=CF_NEGATIVE_FILL)
    )

    # Accroissement Financier (R): same
    ws.conditional_formatting.add(
        f'R{DATA_ROW_START}:R{tr - 1}',
        CellIsRule(operator='greaterThan', formula=['0'],
                   font=CF_POSITIVE_FONT, fill=CF_POSITIVE_FILL)
    )
    ws.conditional_formatting.add(
        f'R{DATA_ROW_START}:R{tr - 1}',
        CellIsRule(operator='lessThan', formula=['0'],
                   font=CF_NEGATIVE_FONT, fill=CF_NEGATIVE_FILL)
    )

    # Écart Physique-Financier (S): green if physical > financial, red if financial > physical
    ws.conditional_formatting.add(
        f'S{DATA_ROW_START}:S{tr - 1}',
        CellIsRule(operator='greaterThan', formula=['0.05'],
                   font=CF_POSITIVE_FONT, fill=CF_POSITIVE_FILL)
    )
    ws.conditional_formatting.add(
        f'S{DATA_ROW_START}:S{tr - 1}',
        CellIsRule(operator='lessThan', formula=['-0.05'],
                   font=CF_NEGATIVE_FONT, fill=CF_NEGATIVE_FILL)
    )
    ws.conditional_formatting.add(
        f'S{DATA_ROW_START}:S{tr - 1}',
        CellIsRule(operator='between', formula=['-0.05', '0.05'],
                   font=CF_WARNING_FONT, fill=CF_WARNING_FILL)
    )

    # Data bars on Avancement Physique % (J)
    ws.conditional_formatting.add(
        f'J{DATA_ROW_START}:J{tr - 1}',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                    color=ACCENT_POSITIVE, showValue=True)
    )

    # Data bars on Avancement Financier % (P)
    ws.conditional_formatting.add(
        f'P{DATA_ROW_START}:P{tr - 1}',
        DataBarRule(start_type='num', start_value=0, end_type='num', end_value=1,
                    color=ACCENT_WARNING, showValue=True)
    )

    # ── Data validation (0-100% for percentage inputs) ─────────
    dv_pct = DataValidation(type="decimal", operator="between", formula1=0, formula2=1,
                            allow_blank=True)
    dv_pct.error = "Veuillez entrer un pourcentage entre 0% et 100%"
    dv_pct.errorTitle = "Valeur invalide"
    dv_pct.prompt = "Entrez le pourcentage (ex: 0.25 pour 25%)"
    dv_pct.promptTitle = "Avancement %"
    ws.add_data_validation(dv_pct)
    dv_pct.add(f'J{DATA_ROW_START}:J{tr - 1}')
    dv_pct.add(f'P{DATA_ROW_START}:P{tr - 1}')

    # ── Column widths ──────────────────────────────────────────
    for c, w in WEEKLY_WIDTHS.items():
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── Freeze panes ───────────────────────────────────────────
    ws.freeze_panes = "C5"

    # ── Protection ─────────────────────────────────────────────
    # Lock reference cells and auto-calculated cells; unlock input cells
    ws.protection.sheet = True
    ws.protection.password = 'ormvag2026'
    for i in range(N):
        r = DATA_ROW_START + i
        for c in REF_COLS + AUTO_COLS:
            ws.cell(row=r, column=c).protection = Protection(locked=True)
        for c in INPUT_COLS:
            ws.cell(row=r, column=c).protection = Protection(locked=False)

    # Date cell unlocked
    ws.cell(row=3, column=3).protection = Protection(locked=False)

    # Tab color
    tab_colors = ["4472C4", "ED7D31", "70AD47", "FFC000"]
    ws.sheet_properties.tabColor = tab_colors[week_idx % len(tab_colors)]


# ═══════════════════════════════════════════════════════════════
# SHEET 6: Récapitulatif Hebdomadaire
# ═══════════════════════════════════════════════════════════════
ws_recap = wb.create_sheet(title="Récapitulatif")

setup_sheet(ws_recap, title="Récapitulatif Hebdomadaire - Suivi d'Exécution", last_col=14)

# ── Section 1: KPI Summary ────────────────────────────────────
# Row 4: KPI headers
kpi_labels = ["", "Total Projets", "Coût Total (DH)", "Avancement Physique Moy.",
              "Avancement Financier Moy.", "Montant Décaissé Total (DH)",
              "Montant Payé Total (DH)", "Écart Physique-Financier"]

for col_idx, label in enumerate(kpi_labels, start=2):
    ws_recap.cell(row=4, column=col_idx, value=label)
style_header_row(ws_recap, row_num=4, col_start=2, col_end=9)

# Row 5: Latest week values (from S04)
last_week = WEEK_NAMES[-1]
tr_last = DATA_ROW_START + N  # totals row in weekly sheet

ws_recap.cell(row=5, column=2, value=f"Semaine en cours ({last_week})")
ws_recap.cell(row=5, column=3).value = N
ws_recap.cell(row=5, column=4).value = f"='{last_week}'!I{tr_last}"
ws_recap.cell(row=5, column=4).number_format = '#,##0'
ws_recap.cell(row=5, column=5).value = f"='{last_week}'!J{tr_last}"
ws_recap.cell(row=5, column=5).number_format = '0.0%'
ws_recap.cell(row=5, column=6).value = f"='{last_week}'!P{tr_last}"
ws_recap.cell(row=5, column=6).number_format = '0.0%'
ws_recap.cell(row=5, column=7).value = f"='{last_week}'!M{tr_last}"
ws_recap.cell(row=5, column=7).number_format = '#,##0'
ws_recap.cell(row=5, column=8).value = f"='{last_week}'!N{tr_last}"
ws_recap.cell(row=5, column=8).number_format = '#,##0'
ws_recap.cell(row=5, column=9).value = f"='{last_week}'!S{tr_last}"
ws_recap.cell(row=5, column=9).number_format = '0.0%'

for col in range(2, 10):
    ws_recap.cell(row=5, column=col).font = font_kpi() if col > 2 else font_subheader()
    ws_recap.cell(row=5, column=col).alignment = align_number() if col > 2 else align_text()

# ── Section 2: Weekly Trend Table ─────────────────────────────
trend_start_row = 8
ws_recap.cell(row=trend_start_row, column=2, value="Évolution Hebdomadaire par Semaine")
ws_recap.cell(row=trend_start_row, column=2).font = font_subheader()
ws_recap.merge_cells(start_row=trend_start_row, start_column=2, end_row=trend_start_row, end_column=9)

trend_header_row = trend_start_row + 1
trend_headers = ["Semaine", "Avancement Physique %", "Accroissement Physique %",
                 "Avancement Financier %", "Accroissement Financier %",
                 "Montant Décaissé (DH)", "Montant Payé (DH)", "Écart Phys.-Fin."]

for col_idx, h in enumerate(trend_headers, start=2):
    ws_recap.cell(row=trend_header_row, column=col_idx, value=h)
style_header_row(ws_recap, row_num=trend_header_row, col_start=2, col_end=9)

for wi, wn in enumerate(WEEK_NAMES):
    r = trend_header_row + 1 + wi
    tr = DATA_ROW_START + N  # totals row in weekly sheet
    ws_recap.cell(row=r, column=2, value=wn)
    ws_recap.cell(row=r, column=3).value = f"='{wn}'!J{tr}"
    ws_recap.cell(row=r, column=3).number_format = '0.0%'
    ws_recap.cell(row=r, column=4).value = f"='{wn}'!L{tr}"
    ws_recap.cell(row=r, column=4).number_format = '0.0%'
    ws_recap.cell(row=r, column=5).value = f"='{wn}'!P{tr}"
    ws_recap.cell(row=r, column=5).number_format = '0.0%'
    ws_recap.cell(row=r, column=6).value = f"='{wn}'!R{tr}"
    ws_recap.cell(row=r, column=6).number_format = '0.0%'
    ws_recap.cell(row=r, column=7).value = f"='{wn}'!M{tr}"
    ws_recap.cell(row=r, column=7).number_format = '#,##0'
    ws_recap.cell(row=r, column=8).value = f"='{wn}'!N{tr}"
    ws_recap.cell(row=r, column=8).number_format = '#,##0'
    ws_recap.cell(row=r, column=9).value = f"='{wn}'!S{tr}"
    ws_recap.cell(row=r, column=9).number_format = '0.0%'

    style_data_row(ws_recap, row_num=r, col_start=2, col_end=9, row_index=wi)

    # Cross-sheet ref color
    for col in range(3, 10):
        ws_recap.cell(row=r, column=col).font = Font(name=FONT_NAME, size=11, color="008000")

# Conditional formatting on trend Accroissement
trend_data_start = trend_header_row + 1
trend_data_end = trend_header_row + NUM_WEEKS
ws_recap.conditional_formatting.add(
    f'D{trend_data_start}:D{trend_data_end}',
    CellIsRule(operator='greaterThan', formula=['0'],
               font=CF_POSITIVE_FONT, fill=CF_POSITIVE_FILL)
)
ws_recap.conditional_formatting.add(
    f'D{trend_data_start}:D{trend_data_end}',
    CellIsRule(operator='lessThan', formula=['0'],
               font=CF_NEGATIVE_FONT, fill=CF_NEGATIVE_FILL)
)
ws_recap.conditional_formatting.add(
    f'F{trend_data_start}:F{trend_data_end}',
    CellIsRule(operator='greaterThan', formula=['0'],
               font=CF_POSITIVE_FONT, fill=CF_POSITIVE_FILL)
)
ws_recap.conditional_formatting.add(
    f'F{trend_data_start}:F{trend_data_end}',
    CellIsRule(operator='lessThan', formula=['0'],
               font=CF_NEGATIVE_FONT, fill=CF_NEGATIVE_FILL)
)

# ── Section 3: Province Summary ───────────────────────────────
prov_start_row = trend_data_end + 3
ws_recap.cell(row=prov_start_row, column=2, value="Récapitulatif par Province")
ws_recap.cell(row=prov_start_row, column=2).font = font_subheader()
ws_recap.merge_cells(start_row=prov_start_row, start_column=2, end_row=prov_start_row, end_column=9)

prov_header_row = prov_start_row + 1
prov_headers = ["Province", "Nb Projets", "Coût Total (DH)", "Avancement Physique %",
                "Avancement Financier %", "Montant Décaissé (DH)", "Montant Payé (DH)", "Écart Phys.-Fin."]

for col_idx, h in enumerate(prov_headers, start=2):
    ws_recap.cell(row=prov_header_row, column=col_idx, value=h)
style_header_row(ws_recap, row_num=prov_header_row, col_start=2, col_end=9)

provinces = ["Kénitra", "Sidi Kacem", "Sidi Slimane"]
for pi, prov in enumerate(provinces):
    r = prov_header_row + 1 + pi

    # Get project indices for this province
    prov_indices = [i for i, p in enumerate(projects) if p["province"] == prov]
    prov_rows = [DATA_ROW_START + idx for idx in prov_indices]

    ws_recap.cell(row=r, column=2, value=prov)
    ws_recap.cell(row=r, column=2).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PROVINCE_COLORS.get(prov, NEUTRAL_900))
    ws_recap.cell(row=r, column=2).fill = PROVINCE_FILLS.get(prov, PatternFill())

    ws_recap.cell(row=r, column=3, value=len(prov_indices))

    # Coût total: sum from Liste Projets
    ws_recap.cell(row=r, column=4).value = f"='Liste Projets'!I{DATA_ROW_START + prov_indices[0]}" if len(prov_indices) == 1 else f"=SUMPRODUCT(('Liste Projets'!C{DATA_ROW_START}:C{DATA_ROW_START + N - 1}=\"{prov}\")*'Liste Projets'!I{DATA_ROW_START}:I{DATA_ROW_START + N - 1})"
    ws_recap.cell(row=r, column=4).number_format = '#,##0'

    # Use SUMPRODUCT to get province-specific averages from last week sheet
    # Avancement Physique % (weighted by cost for this province)
    ws_recap.cell(row=r, column=5).value = f'=IFERROR(SUMPRODUCT((\'{last_week}\'!C{DATA_ROW_START}:C{DATA_ROW_START + N - 1}="{prov}")*\'{last_week}\'!J{DATA_ROW_START}:J{DATA_ROW_START + N - 1}*\'{last_week}\'!I{DATA_ROW_START}:I{DATA_ROW_START + N - 1})/SUMPRODUCT((\'{last_week}\'!C{DATA_ROW_START}:C{DATA_ROW_START + N - 1}="{prov}")*\'{last_week}\'!I{DATA_ROW_START}:I{DATA_ROW_START + N - 1}),0)'
    ws_recap.cell(row=r, column=5).number_format = '0.0%'

    # Avancement Financier % (weighted by cost)
    ws_recap.cell(row=r, column=6).value = f'=IFERROR(SUMPRODUCT((\'{last_week}\'!C{DATA_ROW_START}:C{DATA_ROW_START + N - 1}="{prov}")*\'{last_week}\'!P{DATA_ROW_START}:P{DATA_ROW_START + N - 1}*\'{last_week}\'!I{DATA_ROW_START}:I{DATA_ROW_START + N - 1})/SUMPRODUCT((\'{last_week}\'!C{DATA_ROW_START}:C{DATA_ROW_START + N - 1}="{prov}")*\'{last_week}\'!I{DATA_ROW_START}:I{DATA_ROW_START + N - 1}),0)'
    ws_recap.cell(row=r, column=6).number_format = '0.0%'

    # Montant Décaissé
    ws_recap.cell(row=r, column=7).value = f'=SUMPRODUCT((\'{last_week}\'!C{DATA_ROW_START}:C{DATA_ROW_START + N - 1}="{prov}")*\'{last_week}\'!M{DATA_ROW_START}:M{DATA_ROW_START + N - 1})'
    ws_recap.cell(row=r, column=7).number_format = '#,##0'

    # Montant Payé
    ws_recap.cell(row=r, column=8).value = f'=SUMPRODUCT((\'{last_week}\'!C{DATA_ROW_START}:C{DATA_ROW_START + N - 1}="{prov}")*\'{last_week}\'!N{DATA_ROW_START}:N{DATA_ROW_START + N - 1})'
    ws_recap.cell(row=r, column=8).number_format = '#,##0'

    # Écart
    ws_recap.cell(row=r, column=9).value = f"=E{r}-F{r}"
    ws_recap.cell(row=r, column=9).number_format = '0.0%'

    style_data_row(ws_recap, row_num=r, col_start=2, col_end=9, row_index=pi)
    # Re-apply province styling
    ws_recap.cell(row=r, column=2).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PROVINCE_COLORS.get(prov, NEUTRAL_900))
    ws_recap.cell(row=r, column=2).fill = PROVINCE_FILLS.get(prov, PatternFill())
    for col in range(4, 10):
        ws_recap.cell(row=r, column=col).font = Font(name=FONT_NAME, size=11, color="008000")

# Province totals
prov_total_row = prov_header_row + 1 + len(provinces)
ws_recap.cell(row=prov_total_row, column=2, value="TOTAL")
ws_recap.cell(row=prov_total_row, column=3).value = f"=SUM(C{prov_header_row + 1}:C{prov_total_row - 1})"
ws_recap.cell(row=prov_total_row, column=4).value = f"=SUM(D{prov_header_row + 1}:D{prov_total_row - 1})"
ws_recap.cell(row=prov_total_row, column=4).number_format = '#,##0'
ws_recap.cell(row=prov_total_row, column=5).value = f"='{last_week}'!J{tr_last}"
ws_recap.cell(row=prov_total_row, column=5).number_format = '0.0%'
ws_recap.cell(row=prov_total_row, column=6).value = f"='{last_week}'!P{tr_last}"
ws_recap.cell(row=prov_total_row, column=6).number_format = '0.0%'
ws_recap.cell(row=prov_total_row, column=7).value = f"=SUM(G{prov_header_row + 1}:G{prov_total_row - 1})"
ws_recap.cell(row=prov_total_row, column=7).number_format = '#,##0'
ws_recap.cell(row=prov_total_row, column=8).value = f"=SUM(H{prov_header_row + 1}:H{prov_total_row - 1})"
ws_recap.cell(row=prov_total_row, column=8).number_format = '#,##0'
ws_recap.cell(row=prov_total_row, column=9).value = f"=E{prov_total_row}-F{prov_total_row}"
ws_recap.cell(row=prov_total_row, column=9).number_format = '0.0%'
style_total_row(ws_recap, row_num=prov_total_row, col_start=2, col_end=9)

# Column widths for Récap
RECAP_WIDTHS = {2: 24, 3: 12, 4: 18, 5: 18, 6: 18, 7: 20, 8: 20, 9: 16}
for c, w in RECAP_WIDTHS.items():
    ws_recap.column_dimensions[get_column_letter(c)].width = w

ws_recap.freeze_panes = "C5"
ws_recap.sheet_properties.tabColor = "7030A0"  # Purple tab


# ═══════════════════════════════════════════════════════════════
# SHEET 7: Guide d'utilisation
# ═══════════════════════════════════════════════════════════════
ws_guide = wb.create_sheet(title="Guide")
setup_sheet(ws_guide, title="Guide d'Utilisation du Canevas Hebdomadaire", last_col=8)

guide_content = [
    ["", "", "", "", "", "", "", ""],
    ["RUBRIQUE", "", "DESCRIPTION", "", "", "", "", ""],
    ["Code Couleur", "", "Jaune = Cellule de saisie (modifiable) | Vert = Formule auto (réf. inter-feuille) | Noir = Formule calculée | Bleu = Valeur saisie", "", "", "", "", ""],
    ["Feuilles S01-S04", "", "Chaque feuille représente une semaine. Saisissez uniquement dans les cellules jaunes : Avancement Physique %, Montants, Avancement Financier %, Observations.", "", "", "", "", ""],
    ["Colonnes Automatiques", "", "% Sem. Préc. : récupéré automatiquement de la semaine précédente | Accroissement % : calculé automatiquement | Écart Phys.-Fin. : différence entre avancement physique et financier", "", "", "", "", ""],
    ["Ajouter Semaines", "", "Pour ajouter S05 : Dupliquez S04, renommez en S05, puis modifiez les formules de la colonne K (% Sem. Préc.) pour référencer S04 au lieu de S03.", "", "", "", "", ""],
    ["Protection", "", "Les cellules de référence et les formules sont verrouillées. Mot de passe de déprotection : ormvag2026", "", "", "", "", ""],
    ["Récapitulatif", "", "La feuille Récapitulatif se met à jour automatiquement à partir de la dernière semaine (S04) et affiche les KPI, l'évolution hebdomadaire et le résumé par province.", "", "", "", "", ""],
    ["Format Pourcentages", "", "Saisir 0.25 pour 25%, 0.50 pour 50%, 1.00 pour 100%. Les valeurs doivent être entre 0 et 1.", "", "", "", "", ""],
]

for i, row_data in enumerate(guide_content):
    r = 4 + i
    for j, val in enumerate(row_data):
        ws_guide.cell(row=r, column=2 + j, value=val)

# Style guide headers
ws_guide.cell(row=5, column=2).font = font_subheader()
ws_guide.cell(row=5, column=4).font = font_subheader()

for r in range(6, 4 + len(guide_content)):
    ws_guide.cell(row=r, column=2).font = Font(name=FONT_NAME, size=11, bold=HEADER_BOLD, color=PRIMARY)
    ws_guide.cell(row=r, column=4).font = font_body()
    ws_guide.cell(row=r, column=4).alignment = Alignment(wrap_text=True)

ws_guide.column_dimensions['B'].width = 24
ws_guide.column_dimensions['D'].width = 80
ws_guide.sheet_properties.tabColor = NEUTRAL_600


# ═══════════════════════════════════════════════════════════════
# Move Guide to the beginning (after Liste Projets)
# ═══════════════════════════════════════════════════════════════
# Reorder sheets: Liste Projets, Guide, S01, S02, S03, S04, Récapitulatif
desired_order = ["Liste Projets", "Guide"] + WEEK_NAMES + ["Récapitulatif"]
sheet_map = {ws.title: ws for ws in wb.worksheets}
wb._sheets = [sheet_map[name] for name in desired_order]


# ═══════════════════════════════════════════════════════════════
# Save
# ═══════════════════════════════════════════════════════════════
OUTPUT_PATH = "/home/z/my-project/download/Canevas_Hebdomadaire_ORMVAG_Gharb.xlsx"
wb.save(OUTPUT_PATH)
print(f"Workbook saved to: {OUTPUT_PATH}")
print(f"Projects: {N}")
print(f"Weekly sheets: {', '.join(WEEK_NAMES)}")
print(f"Sheets: {[ws.title for ws in wb.worksheets]}")
