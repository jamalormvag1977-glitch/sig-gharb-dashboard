import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import copy
import shutil
import os

# ============================================================
# Define June-December 2026 weeks (ISO)
# ============================================================
def iso_week_start(year, week):
    """Get Monday of ISO week `week` in `year`."""
    jan4 = datetime(year, 1, 4)
    jan4_day = jan4.isoweekday()  # 1=Mon..7=Sun
    monday_w1 = jan4 - timedelta(days=jan4_day - 1)
    return monday_w1 + timedelta(weeks=week - 1)

JUIN_1 = datetime(2026, 6, 1)

WEEKS_JUIN_DEC = []
for w in range(1, 54):
    monday = iso_week_start(2026, w)
    if monday.year > 2026:
        break
    # Only include weeks whose Monday is on or after June 1, 2026
    if monday < JUIN_1:
        continue
    sunday = monday + timedelta(days=6)
    month = monday.month
    MONTH_NAMES = {1:'Janvier', 2:'Février', 3:'Mars', 4:'Avril', 5:'Mai',
                   6:'Juin', 7:'Juillet', 8:'Août', 9:'Septembre',
                   10:'Octobre', 11:'Novembre', 12:'Décembre'}
    WEEKS_JUIN_DEC.append({
        'iso_week': w,
        'monday': monday,
        'sunday': sunday,
        'month': month,
        'label': f"S{w} ({monday.strftime('%d/%m')} - {sunday.strftime('%d/%m')})",
        'month_name': MONTH_NAMES.get(month, '')
    })

print(f"Total weeks Juin-Déc: {len(WEEKS_JUIN_DEC)}")
for wk in WEEKS_JUIN_DEC[:3]:
    print(f"  S{wk['iso_week']}: {wk['monday'].strftime('%d/%m/%Y')} - {wk['sunday'].strftime('%d/%m/%Y')} ({wk['month_name']})")
print("  ...")
for wk in WEEKS_JUIN_DEC[-3:]:
    print(f"  S{wk['iso_week']}: {wk['monday'].strftime('%d/%m/%Y')} - {wk['sunday'].strftime('%d/%m/%Y')} ({wk['month_name']})")

MONTHS_JUIN_DEC = [
    (6, "Juin 2026"),
    (7, "Juillet 2026"),
    (8, "Août 2026"),
    (9, "Septembre 2026"),
    (10, "Octobre 2026"),
    (11, "Novembre 2026"),
    (12, "Décembre 2026"),
]

# ============================================================
# 1. UPDATE Suivi_Hebdomadaire_Gharb_2026.xlsx
# ============================================================
print("\n=== Updating Suivi_Hebdomadaire_Gharb_2026.xlsx ===")
src = 'public/Suivi_Hebdomadaire_Gharb_2026.xlsx'
dst = 'public/Suivi_Hebdomadaire_Gharb_2026.xlsx'
wb = openpyxl.load_workbook(src)

# Sheet 1: Suivi Hebdomadaire - update header
ws = wb['Suivi Hebdomadaire']
ws['B2'] = "SUIVI HEBDOMADAIRE D'EXÉCUTION PHYSIQUE ET FINANCIÈRE — PROJETS INONDATIONS GHARB 2026 (Juin–Décembre)"
ws['B3'] = 'Office Régional de Mise en Valeur Agricole du Gharb — Kénitra'

# Sheet 2: Historique Hebdomadaire - keep only weeks 23-52/53
ws_hist = wb['Historique Hebdomadaire']
ws_hist['B2'] = "HISTORIQUE D'AVANCEMENT PAR SEMAINE — VUE D'ENSEMBLE (Juin–Décembre 2026)"
ws_hist['B3'] = "Ce tableau permet de suivre l'évolution semaine par semaine de l'avancement global par province (Juin à Décembre 2026)"

# Clear existing week rows
for row_idx in range(6, ws_hist.max_row + 1):
    for col_idx in range(2, 13):
        ws_hist.cell(row=row_idx, column=col_idx).value = None

# Write new week rows (starting at row 6)
row_idx = 6
for wk in WEEKS_JUIN_DEC:
    ws_hist.cell(row=row_idx, column=2, value=wk['iso_week'])
    ws_hist.cell(row=row_idx, column=3, value=f"{wk['monday'].strftime('%d/%m/%Y')} - {wk['sunday'].strftime('%d/%m/%Y')}")
    row_idx += 1

# Sheet 3: Récap. par Province - update header
ws_recap = wb['Récap. par Province']
ws_recap['B2'] = "RÉCAPITULATIF PAR PROVINCE — SEMAINE EN COURS (Juin–Décembre 2026)"

# Sheet 4: Mode d'emploi - update
ws_guide = wb["Mode d'emploi"]
for row in ws_guide.iter_rows(min_row=1, max_row=ws_guide.max_row, values_only=False):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and '72 projets' in cell.value:
            cell.value = "Ce fichier permet d'assurer un suivi hebdomadaire (Juin à Décembre 2026) de l'exécution physique et financière des 72 projets de lutte contre les inondations dans la région du Gharb."

wb.save(dst)
print("  Saved Suivi_Hebdomadaire_Gharb_2026.xlsx")
shutil.copy2(dst, 'download/Suivi_Hebdomadaire_Gharb_2026.xlsx')

# ============================================================
# 2. UPDATE Modele_Suivi_Avancement_Gharb_2026.xlsx
# ============================================================
print("\n=== Updating Modele_Suivi_Avancement_Gharb_2026.xlsx ===")
src = 'public/Modele_Suivi_Avancement_Gharb_2026.xlsx'
dst = 'public/Modele_Suivi_Avancement_Gharb_2026.xlsx'
wb = openpyxl.load_workbook(src)

# Sheet 1: Suivi des Projets - update header
ws = wb['Suivi des Projets']
ws['B2'] = 'Suivi Physique et Financier des Projets — Région du Gharb 2026 (Juin–Décembre)'

# Sheet 2: Récap. par Province - update header
ws = wb['Récap. par Province']
ws['B2'] = 'Récapitulatif par Province — Région du Gharb 2026 (Juin–Décembre)'

# Sheet 3: Récap. par Secteur - update header
ws = wb['Récap. par Secteur']
ws['B2'] = 'Récapitulatif par Secteur — Région du Gharb 2026 (Juin–Décembre)'

# Sheet 4: Mode d'emploi
ws = wb["Mode d'emploi"]
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
    for cell in row:
        if cell.value and isinstance(cell.value, str) and 'dashboard' in cell.value.lower():
            cell.value = cell.value.replace(
                'avec les données de suivi physique et financier des projets.',
                'avec les données de suivi physique et financier des projets (Juin à Décembre 2026).'
            )

wb.save(dst)
print("  Saved Modele_Suivi_Avancement_Gharb_2026.xlsx")
shutil.copy2(dst, 'download/Modele_Suivi_Avancement_Gharb_2026.xlsx')

# ============================================================
# 3. REBUILD Canevas_Hebdomadaire_ORMVAG_Gharb.xlsx
#    Keep only S23-S53, update all references
# ============================================================
print("\n=== Rebuilding Canevas_Hebdomadaire_ORMVAG_Gharb.xlsx ===")
src = 'public/Canevas_Hebdomadaire_ORMVAG_Gharb.xlsx'
dst = 'public/Canevas_Hebdomadaire_ORMVAG_Gharb.xlsx'
wb = openpyxl.load_workbook(src)

# Get the list of weekly sheet names to remove (S01-S22)
sheets_to_remove = [f"S{i:02d}" for i in range(1, 23)]
for name in sheets_to_remove:
    if name in wb.sheetnames:
        del wb[name]
        print(f"  Removed sheet: {name}")

# Update each remaining weekly sheet (S23-S52/S53)
for name in wb.sheetnames:
    if name.startswith('S') and name[1:].isdigit():
        ws = wb[name]
        week_num = int(name[1:])
        # Find the week info
        wk_info = None
        for wk in WEEKS_JUIN_DEC:
            if wk['iso_week'] == week_num:
                wk_info = wk
                break
        if wk_info:
            # Update the title in row 2
            new_title = f"Suivi Hebdomadaire - Semaine {week_num} ({wk_info['monday'].strftime('%d/%m')} - {wk_info['sunday'].strftime('%d/%m')}) — {wk_info['month_name']} 2026 - Exécution Physique et Financière"
            ws.cell(row=2, column=2).value = new_title
            # Update the date in row 3
            ws.cell(row=3, column=3).value = wk_info['monday']
            print(f"  Updated sheet {name}: {wk_info['month_name']} 2026")

# Determine the last weekly sheet name
last_week_sheet = None
for name in reversed(wb.sheetnames):
    if name.startswith('S') and name[1:].isdigit():
        last_week_sheet = name
        break
print(f"  Last weekly sheet: {last_week_sheet}")

# Update Récapitulatif sheet
if 'Récapitulatif' in wb.sheetnames:
    ws_recap = wb['Récapitulatif']
    ws_recap.cell(row=2, column=2).value = "Récapitulatif Hebdomadaire - Suivi d'Exécution 2026 (Juin–Décembre)"
    # Update formulas to reference last week sheet
    if last_week_sheet:
        ws_recap.cell(row=4, column=3).value = f"='{last_week_sheet}'!I77"  # Coût Total
        ws_recap.cell(row=5, column=3).value = f"='{last_week_sheet}'!N77"  # Avancement Physique Moy.
        ws_recap.cell(row=6, column=3).value = f"='{last_week_sheet}'!T77"  # Avancement Financier Moy.
        ws_recap.cell(row=7, column=3).value = f"='{last_week_sheet}'!Q77"  # Montant Décaissé Total
        ws_recap.cell(row=8, column=3).value = f"='{last_week_sheet}'!R77"  # Montant Payé Total

# Update Guide sheet
if 'Guide' in wb.sheetnames:
    ws_guide = wb['Guide']
    ws_guide.cell(row=2, column=2).value = "Guide d'Utilisation du Canevas Hebdomadaire (Juin–Décembre 2026)"
    for row in ws_guide.iter_rows(min_row=1, max_row=ws_guide.max_row, values_only=False):
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                if 'S01-S52' in cell.value:
                    cell.value = cell.value.replace('S01-S52', 'S23-S53')
                if '52 feuilles' in cell.value:
                    cell.value = cell.value.replace('52 feuilles', f'{len(WEEKS_JUIN_DEC)} feuilles')
                if 'semaine ISO' in cell.value.lower():
                    cell.value = cell.value.replace('semaine ISO', 'semaine ISO (S23=S1er Juin)')

# Update Liste Projets header
if 'Liste Projets' in wb.sheetnames:
    ws_proj = wb['Liste Projets']
    ws_proj.cell(row=2, column=2).value = 'Liste Référentielle des Projets - Indicateurs Physiques (Juin–Décembre 2026)'

wb.save(dst)
print("  Saved Canevas_Hebdomadaire_ORMVAG_Gharb.xlsx")
shutil.copy2(dst, 'download/Canevas_Hebdomadaire_ORMVAG_Gharb.xlsx')

print("\n=== All Excel files updated for Juin–Décembre 2026 ===")
